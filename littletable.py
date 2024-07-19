#
#
# littletable.py
#
# littletable is a simple in-memory database for ad-hoc or user-defined objects,
# supporting simple query and join operations - useful for ORM-like access
# to a collection of data objects, without dealing with SQL
#
#
# Copyright (c) 2010-2023  Paul T. McGuire
#
# Permission is hereby granted, free of charge, to any person obtaining
# a copy of this software and associated documentation files (the
# "Software"), to deal in the Software without restriction, including
# without limitation the rights to use, copy, modify, merge, publish,
# distribute, sublicense, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to
# the following conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
# MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY
# CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
# TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
# SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
#
from __future__ import annotations

__doc__ = r"""

C{littletable} - a Python module to give ORM-like access to a collection of objects

The C{littletable} module provides a low-overhead, schema-less, in-memory database access to a 
collection of user objects.  C{Table}s can
contain user-defined objects, using those objects' C{__dict__}, C{__slots__}, or C{_fields}
mappings to access object attributes. Table contents can thus also include namedtuples, 
SimpleNamespaces, or dataclasses.

C{Tables} can also be constructed using Python dicts. In this case, they are stored as 
SimpleNamespaces.

In addition to basic insert/remove/query/delete access to the contents of a 
Table, C{littletable} offers:
 - indexing for improved retrieval performance, and optional enforcing key uniqueness
 - access to objects using indexed attributes
 - full text search on attributes containing extended text content
 - simplified joins using '+' operator syntax between annotated Tables
 - the result of any query or join is a new first-class C{littletable} Table
 - pivot on one or two attributes to gather tabulated data summaries
 - easy import/export to CSV and JSON files

C{littletable} Tables do not require an upfront schema definition, but simply work off of the
attributes in the stored values, and those referenced in any query parameters.

Here is a simple C{littletable} data storage/retrieval example::

    from littletable import Table

    # create table of customers
    customers = Table('customers')
    customers.create_index("id", unique=True)
    customers.insert({id="0010", name="George Jetson"})
    customers.insert({id="0020", name="Wile E. Coyote"})
    customers.insert({id="0030", name="Jonny Quest"})

    # create table of product catalog (load from CSV data)
    catalog_data = '''\
    sku,descr,unitofmeas,unitprice
    BRDSD-001,Bird seed,LB,3
    BBS-001,Steel BB's,LB,5
    MAGNT-001,Magnet,EA,8
    MAGLS-001,Magnifying glass,EA,12
    ANVIL-001,1000lb anvil,EA,100
    ROPE-001,1 in. heavy rope,100FT,10
    ROBOT-001,Domestic robot,EA,5000'''

    catalog = lt.Table("catalog")
    catalog.create_index("sku", unique=True)
    catalog.csv_import(catalog_data, transforms={'unitprice':int})

    print(catalog.by.sku["ANVIL-001"].descr)

    # create many-to-many link table of wishlist items
    wishitems = Table('wishitems')
    wishitems.create_index("custid")
    wishitems.create_index("sku")
    wishitems.insert({custid="0020", sku="ANVIL-001"})
    wishitems.insert({custid="0020", sku="BRDSD-001"})
    wishitems.insert({custid="0020", sku="MAGNT-001"})
    wishitems.insert({custid="0030", sku="MAGNT-001"})
    wishitems.insert({custid="0030", sku="MAGLS-001"})

    # print a particular customer name 
    # (unique indexes will return a single item; non-unique
    # indexes will return a list of all matching items)
    print(customers.by.id["0030"].name)

    # print all items sold by the pound
    for item in catalog.where(unitofmeas="LB"):
        print(item.sku, item.descr)

    # print all items that cost more than 10
    for item in catalog.where(lambda o : o.unitprice>10):
        print(item.sku, item.descr, item.unitprice)

    # join tables to create queryable wishlists collection
    wishlists = customers.join_on("id") + wishitems.join_on("custid") + catalog.join_on("sku")

    # print all wishlist items with price > 10
    bigticketitems = wishlists().where(lambda ob : ob.unitprice > 10)
    for item in bigticketitems:
        print(item)

    # list all wishlist items in descending order by price
    for item in wishlists().orderby("unitprice desc"):
        print(item)
"""

import base64
import contextlib
import copy
import csv
import datetime
import functools
import io
import itertools
import textwrap
import warnings
from enum import Enum
from io import StringIO
import json
import operator
import os
import random
import re
import shlex
import statistics
import sys
from collections import defaultdict, namedtuple, Counter
from collections.abc import Mapping, Sequence
from functools import partial
from pathlib import Path
from types import SimpleNamespace
import urllib.request
from typing import (
    Callable, Any, TextIO, Union, Optional, Iterable, Iterator,
    Generic, TypeVar, Type, cast, Tuple, List,
)

try:
    import rich
    from rich import box
except ImportError:
    rich = None
    box = None

version_info = namedtuple("version_info", "major minor micro release_level serial")
__version_info__ = version_info(2, 3, 3, "final", 0)
__version__ = (
    "{}.{}.{}".format(*__version_info__[:3])
    + (f"{__version_info__.release_level[0]}{__version_info__.serial}", "")[
        __version_info__.release_level == "final"
    ]
)
__version_time__ = "19 Jul 2024 15:30 UTC"
__author__ = "Paul McGuire <ptmcg@austin.rr.com>"


# custom Exception classes
class SearchIndexInconsistentError(Exception):
    """
    Exception raised when using search method on table that has been
    modified since the search index was built.
    """


class NoSuchIndexError(KeyError):
    """
    Exception raised when trying to access an index that does not exist.
    """
    def __init__(self, attrname: str):
        super().__init__(attrname)

    def __str__(self) -> str:
        index_name = super().__str__()
        return f"no such index {index_name!r}"


class UnableToExtractAttributeNamesError(ValueError):
    """
    Exception raised when attributes cannot be determined from an object.
    """


class ReadonlyIndexAccessError(Exception):
    """
    Exception raised when trying to write to a readonly index.
    """


class AuthenticationWarning(Warning):
    """
    Warning emitted when using authentication credentials with http:// URL.
    """
    def __str__(self) -> str:
        return (
            "Using Basic Authentication over HTTP can expose login credentials; HTTPS is recommended"
        )


def _emit_warning_with_user_frame(warning: Warning) -> None:
    import sys
    import warnings

    try:
        cur = sys._getframe()
    except AttributeError:
        user_stack_level = 2
    else:
        # walk stack trace until outside of this module
        user_stack_level = 0
        while cur:
            user_stack_level += 1
            if cur.f_code.co_filename != __file__:
                break
            cur = cur.f_back
        else:
            user_stack_level = 2

    warnings.warn(message=str(warning), category=type(warning), stacklevel=user_stack_level)


class attrgetter:  # noqa
    """
    Return a callable object that fetches the given attributes(s) from its operand,
    and returns their values as a tuple.

    Accepts an optional C{defaults} dict for any attribute that is not present in the
    given object. If an attribute is not present and no default value is defined for
    that attribute, fills in None for that attribute.

    Not quite a drop-in replacement for C{operator.attrgetter} - that method will accept
    dotted attribute names and will traverse the object path to extract attributes
    from contained objects.
    """
    __slots__ = ('_items', '_defaults', '_call')

    def __init__(self, item: str, /, *items, defaults: dict[str, Any] = None):

        defaults = defaults or {}
        self._defaults = {**defaults}

        if not items:
            # only a single attribute name given
            self._items = (item,)
            default_value = defaults.get(item)

            def func(obj):
                return (getattr(obj, item, default_value),)

        else:
            # multiple attribute names given (first is item, the rest are in items,
            # so we must merge into a single tuple).
            items = (item,) + items
            self._items = items
            base_getter = operator.attrgetter(*items)
            item_default_values = tuple(
                (k, self._defaults.get(k)) for k in items
            )

            def func(obj):
                try:
                    return base_getter(obj)
                except AttributeError:
                    return tuple(
                        getattr(obj, *item_default)
                        for item_default in item_default_values
                    )

        self._call = func

    def __call__(self, obj, /):
        return self._call(obj)

    def __repr__(self) -> str:
        if self._defaults:
            return (
                f"{self.__class__.__module__}.{self.__class__.__name__}"
                f"({', '.join(map(repr, self._items))}, {self._defaults})"
            )
        else:
            return (
                f"{self.__class__.__module__}.{self.__class__.__name__}"
                f"({', '.join(map(repr, self._items))})"
            )

    def __reduce__(self):
        return self.__class__, self._items, self._defaults


NL = os.linesep

default_row_class = SimpleNamespace

_numeric_type: tuple[type, ...] = (int, float)
right_justify_types: tuple[type, ...] = (int, float, datetime.timedelta)

try:
    import numpy
except ImportError:
    numpy = None
else:
    _numeric_type += (numpy.number,)

try:
    import openpyxl
except ImportError:
    openpyxl = None

PredicateFunction = Callable[[Any], bool]

__all__ = [
    "AuthenticationWarning",
    "DataObject",
    "FixedWidthReader",
    "Table",
    "csv_import",
    "tsv_import",
    "json_import",
    "excel_import",
]

# define default stopwords for full_text_search
_stopwords = frozenset(
    (*"""\
     a about above after again against all am an and any are aren't as at be because been 
     before being below between both but by can't cannot could couldn't did didn't do does 
     doesn't doing don't down during each few for from further had hadn't has hasn't have haven't 
     having he he'd he'll he's her here here's hers herself him himself his how how's i i'd i'll
     i'm i've if in into is isn't it it's its itself let's me more most mustn't my myself no nor
     not of off on once only or other ought our ours ourselves out over own same shan't she she'd 
     she'll she's should shouldn't so some such than that that's the their theirs them themselves 
     then there there's these they they'd they'll they're they've this those through to too under 
     until up very was wasn't we we'd we'll we're we've were weren't what what's when when's 
     where where's which while who who's whom why why's with won't would wouldn't you 
     you'd you'll you're you've your yours yourself yourselves""".split(),
     *"d ll m re s t ve".split(), "",
     )
)

# irregular plurals and singulars for text search handling
_common_english_irregular_plurals = {
    'addenda': 'addendum', 'addendums': 'addendum', 'alumnae': 'alumna', 'alumni': 'alumnus', 'analyses': 'analysis',
    'antennae': 'antenna', 'antennas': 'antenna', 'antitheses': 'antithesis',
    'appendices': 'appendix', 'appendixes': 'appendix', 'bacilli': 'bacillus', 'bacteria': 'bacterium',
    'cacti': 'cactus', 'calves': 'calf', 'children': 'child', 'corpora': 'corpus', 'crises': 'crisis',
    'criteria': 'criterion', 'curricula': 'curriculum', 'diagnoses': 'diagnosis', 'dice': 'die',
    'dwarves': 'dwarf', 'dwarfs': 'dwarf', 'elves': 'elf', 'ellipses': 'ellipsis', 'errata': 'erratum',
    'firemen': 'fireman', 'foci': 'focus', 'feet': 'foot', 'formulae': 'formula', 'fungi': 'fungus', 'genera': 'genus',
    'geese': 'goose', 'halves': 'half', 'hooves': 'hoof', 'hypotheses': 'hypothesis',
    'indices': 'index', 'indexes': 'index', 'knives': 'knife', 'larvae': 'larva', 'leaves': 'leaf', 'lives': 'life',
    'loaves': 'loaf', 'loci': 'locus', 'lice': 'louse', 'men': 'man', 'matrices': 'matrix', 'media': 'medium',
    'memoranda': 'memorandum', 'minutiae': 'minutia', 'mice': 'mouse', 'nebulae': 'nebula', 'nuclei': 'nucleus',
    'oases': 'oasis', 'opera': 'opus', 'ova': 'ovum', 'oxen': 'ox', 'parentheses': 'parenthesis',
    'phenomena': 'phenomenon', 'phyla': 'phylum', 'quizzes': 'quiz', 'radii': 'radius', 'referenda': 'referendum',
    'scarves': 'scarf', 'selves': 'self', 'shelves': 'shelf', 'staves': 'staff', 'stimuli': 'stimulus',
    'strata': 'stratum', 'syllabi': 'syllabus', 'symposia': 'symposium', 'synopses': 'synopsis', 'tableaux': 'tableau',
    'theses': 'thesis', 'thieves': 'thief', 'teeth': 'tooth', 'vertebrae': 'vertebra', 'vertices': 'vertex',
    'vitae': 'vita', 'vortices': 'vortex', 'wharves': 'wharf', 'wives': 'wife', 'wolves': 'wolf', 'women': 'woman',
}
_singulars_that_look_like_plurals = [
    'rabies', 'scabies', 'caries', 'aries', 'series', 'billiards', 'grits', 'pliers', 'whereabouts', 'jeans',
    'binoculars', 'scissors', 'tidings', 'trousers', 'clothes', 'news', 'measles', 'mumps', 'calculus', 'molasses',
    'tweezers', 'dominoes', 'pants', 'odds', 'riches', 'alms', 'barracks', 'chassis', 'corps', 'headquarters', 'ides',
    'kudos', 'species'
]
_plurals_map = {
    **_common_english_irregular_plurals,
    **{s: s for s in _singulars_that_look_like_plurals}
}

_significant_word_endings = (
    'error',
    'warning',
    'exception',
)


def _object_attrnames(obj: Any) -> list[str]:
    if hasattr(obj, "trait_names"):
        return obj.trait_names()
    if hasattr(obj, "__dict__"):
        # normal object
        return list(obj.__dict__)
    if isinstance(obj, tuple) and hasattr(obj, "_fields"):
        # namedtuple
        return obj._fields
    if hasattr(obj, "__slots__"):
        return list(obj.__slots__)
    raise UnableToExtractAttributeNamesError(
        f"object of type {type(obj).__name__!r} has unknown attributes"
    )


def _to_dict(obj: Any) -> dict[str, Any]:
    if hasattr(obj, "trait_names"):
        return {
            k: v
            for k, v in zip(obj.trait_names(), (getattr(obj, a) for a in obj.trait_names()))
        }
    if hasattr(obj, "__dict__"):
        # normal object
        return obj.__dict__
    if isinstance(obj, tuple) and hasattr(obj, "_fields"):
        # namedtuple
        return dict(zip(obj._fields, obj))
    if hasattr(obj, "__slots__"):
        return {
            k: v
            for k, v in zip(obj.__slots__, (getattr(obj, a) for a in obj.__slots__))
        }
    raise UnableToExtractAttributeNamesError(
        f"object of type {type(obj).__name__!r} has unknown attributes"
    )


def _to_json(obj, enc_cls: Type[json.JSONEncoder], **kwargs: Any) -> str:
    return json.dumps(_to_dict(obj), cls=enc_cls, **kwargs)


class DataObject:
    """
    A generic semi-mutable object for storing data values in a table. Attributes
    can be set by passing in named arguments in the constructor, or by setting them
    as C{object.attribute = value}. New attributes can be added any time, but updates
    are ignored.  Table joins are returned as a Table of DataObjects.
    """

    def __init__(self, **kwargs: Any):
        warnings.warn(
            "littletable.DataObject class is deprecated, use types.Simplenamespace or Python dict",
            DeprecationWarning,
            stacklevel=2,
        )
        if kwargs:
            self.__dict__.update(kwargs)

    def __repr__(self):
        return (
            "{"
            f"""{', '.join(f"{k!r}: {v!r}" for k, v in sorted(self.__dict__.items()))}"""
            "}"
        )

    def __setattr__(self, attr, val):
        # make all attributes write-once
        if attr not in self.__dict__:
            super().__setattr__(attr, val)
        else:
            raise AttributeError("can't set existing attribute")

    def __hasattr__(self, key):
        return key in self.__dict__

    def __getitem__(self, k):
        if hasattr(self, k):
            return getattr(self, k)
        else:
            raise KeyError("object has no such attribute " + k)

    __iter__ = None

    def __setitem__(self, k, v):
        if k not in self.__dict__:
            self.__dict__[k] = v
        else:
            raise KeyError("attribute already exists")

    def __eq__(self, other):
        return self.__dict__ == other.__dict__

    def __ne__(self, other):
        return not (self == other)


class _ObjIndex:
    def __init__(self, attr: str):
        self.attr = attr
        self.obs_lookup: defaultdict[str, list] = defaultdict(list)
        self.is_unique = False

    def sort(self, key: Any, reverse: bool = False) -> None:
        for seq in self.obs_lookup.values():
            seq.sort(key=key, reverse=reverse)

    def __setitem__(self, k, v) -> None:
        self.obs_lookup[k].append(v)

    def __getitem__(self, k) -> Any:
        return self.obs_lookup.get(k, [])

    def __len__(self) -> int:
        return len(self.obs_lookup)

    def __iter__(self) -> Iterator[Any]:
        return iter(self.obs_lookup)

    def keys(self) -> list[Any]:
        return sorted(filter(partial(operator.ne, None), self.obs_lookup))

    def items(self) -> Iterable[tuple[Any, Any]]:
        return self.obs_lookup.items()

    def remove(self, obj) -> None:
        try:
            k = getattr(obj, self.attr)
            self.obs_lookup[k].remove(obj)
        except (ValueError, AttributeError, KeyError):
            pass

    def __contains__(self, key: Any) -> bool:
        return key in self.obs_lookup

    def copy_template(self):
        return self.__class__(self.attr)

    def get(self, key, default=None):
        if key in self:
            return self[key]
        else:
            return default

    def _clear(self) -> None:
        self.obs_lookup.clear()


Mapping.register(_ObjIndex)


class _UniqueObjIndex(_ObjIndex):
    def __init__(self, attr, accept_none=False):
        super().__init__(attr)
        self.obs_lookup = {}
        self.is_unique = True
        self.accept_none = accept_none
        self.none_values = []

    def sort(self, key, reverse: bool = False):
        pass

    def __setitem__(self, k, v):
        if k is not None:
            if k not in self.obs_lookup:
                self.obs_lookup[k] = v
            else:
                raise KeyError(f"duplicate key value {k!r}")
        else:
            if self.accept_none:
                self.none_values.append(v)
            else:
                raise ValueError("None is not a valid index key")

    def __getitem__(self, k):
        if k is not None:
            return [self.obs_lookup.get(k)] if k in self.obs_lookup else []
        else:
            return list(self.none_values)

    def __contains__(self, k):
        if k is not None:
            return k in self.obs_lookup
        else:
            return self.accept_none and self.none_values

    def keys(self):
        return sorted(self.obs_lookup) + ([None, ] if self.none_values else [])

    def items(self):
        return ((k, [v]) for k, v in self.obs_lookup.items())

    def remove(self, obj):
        if (k := getattr(obj, self.attr)) is not None:
            self.obs_lookup.pop(k, None)
        else:
            try:
                self.none_values.remove(obj)
            except ValueError:
                pass

    def _clear(self):
        super()._clear()
        del self.none_values[:]


class _ObjIndexWrapper:
    def __init__(self, ind, table):
        self._index = ind
        self._table: Table = table

    def __getattr__(self, attr):
        return getattr(self._index, attr)

    def __call__(self, attr):
        return getattr(self, attr)

    def _getitem_using_slice(self, k):
        where_selector = {
            (False, False, False): lambda: ValueError("must specify start and/or stop values for slice"),
            (False, True, False): lambda: Table.lt(k.stop),
            (True, False, False): lambda: Table.ge(k.start),
            (True, True, False): lambda: (Table.in_range(k.start, k.stop)
                                          if k.start < k.stop
                                          else ValueError("slice end must be greater than slice start")),
            (False, False, True): lambda: ValueError("step slicing not supported"),
            (True, False, True): lambda: ValueError("step slicing not supported"),
            (False, True, True): lambda: ValueError("step slicing not supported"),
            (True, True, True): lambda: ValueError("step slicing not supported"),
        }[k.start is not None,
          k.stop is not None,
          k.step is not None]()

        if isinstance(where_selector, Exception):
            raise where_selector

        return self._table.where(**{self._index.attr: where_selector})

    def __getitem__(self, k):
        if isinstance(k, slice):
            return self._getitem_using_slice(k)

        ret = self._table.copy_template()
        if k in self._index:
            ret.insert_many(self._index[k])
        return ret

    __iter__ = None

    def __contains__(self, k):
        return k in self._index

    def get(self, key, default=None):
        if key in self:
            return self[key]
        else:
            return default


Mapping.register(_ObjIndexWrapper)


class _UniqueObjIndexWrapper(_ObjIndexWrapper):
    def __getitem__(self, k):
        if k is not None:
            if isinstance(k, slice):
                return super().__getitem__(k)
            try:
                return self._index[k][0]
            except IndexError:
                raise KeyError(f"no such value {k!r} in index {self._index.attr!r}")
        else:
            ret = self._table_template.copy_template()
            if k in self._index:
                ret.insert_many(self._index[k])
            return ret


class _ReadonlyObjIndexWrapper(_ObjIndexWrapper):
    def __setitem__(self, k, value):
        raise ReadonlyIndexAccessError(f"no update access to index {self.attr!r}")


class _TableAttributeValueLister:
    class UniquableIterator:
        def __init__(self, seq):
            self.__seq = seq
            self.__iter = iter(seq)

        def __iter__(self):
            return self

        def __next__(self):
            return next(self.__iter)

        def __getattr__(self, attr):
            if attr == "unique":
                self.__iter = filter(
                    lambda x, seen=set(): x not in seen and not seen.add(x), self.__iter  # noqa
                )
                return self
            raise AttributeError(f"no such attribute {attr!r} defined")

    def __init__(self, table, default=None):
        self.__table = table
        self.__default = default

    def __getattr__(self, attr):
        vals = (getattr(row, attr, self.__default) for row in self.__table)
        return _TableAttributeValueLister.UniquableIterator(vals)

    def __call__(self, attr):
        return getattr(self, attr)


class _TableSearcher:
    def __init__(self, table):
        self.__table = table

    def __getattr__(self, attr):
        ret = partial(self.__table._search, attr)
        ret.__name__ = f"Table.search.{attr}"
        ret.__doc__ = """
        {0}
        Search function for attribute {1!r} in a Table.
        
        Parameters:
        - query: str - query expression containing one or more search terms
        - limit: int - limit number of returned values, default=no limit
        - min_score: int - minimum matching score for returned values
        - include_words - include each records raw words in the search results
        - as_table: bool - flag to select whether to return the results as a
          of tuples or as a table (default=False, returns as a list of tuples)
        
        Query search terms can be single words, optionally marked with 
        leading +, ++, -, or -- prefixes to indicate preference in records
        with those words. 
        
        Returns:
        - list or Table of search result tuples, each tuple containing:
          - matched table record
          - search score
          - (if include_words=True) deduplicated list of words in the matched
            record's {1!r}
        """.format(ret.__name__, attr)
        ret.__annotations__ = {
            "query": str,
            "limit": int,
            "min_score": int,
            "include_words": bool,
            "as_table": bool,
            "return": 'Table | list[tuple]'
        }
        return ret

    def __call__(self, attr):
        return getattr(self, attr)

    def __dir__(self):
        return list(self.__table._search_indexes)


class _IndexAccessor:
    def __init__(self, table):
        self._table = table

    def __dir__(self):
        ret = list(self._table._indexes)
        return ret

    def __getattr__(self, attr):
        """
        A quick way to query for matching records using their indexed attributes. The attribute
        name is used to locate the index, and returns a wrapper on the index.  This wrapper provides
        dict-like access to the underlying records in the table, as in::

           employees.by.socsecnum["000-00-0000"]
           customers.by.zipcode["12345"]

        (C{'by'} is added as a pseudo-attribute on tables, to help indicate that the indexed attributes
        are not attributes of the table, but of items in the table.)

        The behavior differs slightly for unique and non-unique indexes:
          - if the index is unique, then retrieving a matching object, will return just the object;
            if there is no matching object, C{KeyError} is raised (making a table with a unique
            index behave very much like a Python dict)
          - if the index is non-unique, then all matching objects will be returned in a new Table,
            just as if a regular query had been performed; if no objects match the key value, an empty
            Table is returned and no exception is raised.

        If there is no index defined for the given attribute, then C{AttributeError} is raised.
        """
        attr_index = self._table._indexes.get(attr)
        if attr_index is not None:
            if isinstance(attr_index, _UniqueObjIndex):
                attr_index_wrapper = _UniqueObjIndexWrapper(attr_index, self._table)
                attr_index_wrapper.__doc__ = textwrap.dedent(
                    f"""\
                    Index accessor by {attr!r}
                    
                    tbl.by.{attr}[attr_value] returns the object having {attr}=attr_value.
                    If no such object exists, raises KeyError.
                    """
                )

            else:
                attr_index_wrapper = _ObjIndexWrapper(attr_index, self._table)
                attr_index_wrapper.__doc__ = textwrap.dedent(
                    f"""\
                    Index accessor by {attr!r}

                    tbl.by.{attr}[attr_value] returns a new Table of all objects having {attr}=attr_value.
                    If no matching objects exist, returns an empty Table.
                    """
                )

            return attr_index_wrapper

        raise AttributeError(f"Table {self._table.table_name!r} has no index {attr!r}")

    def __call__(self, attr):
        return getattr(self, attr)


_ImportExportDataContainer = Union[str, Path, Iterable[str], TextIO]


class ImportSourceType(Enum):
    file = 0
    string = 1
    url = 2
    path = 3
    zip = 4
    gzip = 5
    lzma = 6
    iterable = 7
    tar_gzip = 8


class _MultiIterator(Iterator):
    """
    Internal wrapper class to put a consistent iterable and
    closeable interface on any of the types that might be used
    as import sources:
    - a str containing raw xSV data (denoted by a containing \n)
    - a str URL (denoted by a leading "http")
    - a str file name (containing the data, or a compressed
      file containing data)
    - a path to a file
    - any iterable
    """

    def __init__(
            self,
            seqobj: _ImportExportDataContainer,
            encoding: str = "utf-8",
            url_args: Optional[dict] = None,
    ):
        def _decoder(seq: Iterable[bytes]) -> Iterable[str]:
            for line in seq:
                yield line.decode(encoding)

        self.type = None
        if isinstance(seqobj, Path):
            seqobj = str(seqobj)
            self.type = ImportSourceType.path

        self._iterobj: Iterable[str]
        self._closeobj = None

        if isinstance(seqobj, str):
            if "\n" in seqobj:
                self._iterobj = iter(StringIO(seqobj))
                self.type = ImportSourceType.string
            elif seqobj.startswith(("http://", "https://")):
                url_args = url_args or {}
                urlopen_args = {}

                if not isinstance(url_args.get("data", b""), bytes):
                    raise TypeError("'data' must be of type bytes")

                if "username" in url_args:
                    if seqobj.startswith("http://"):
                        _emit_warning_with_user_frame(AuthenticationWarning())

                    creds = f"{url_args.pop('username')}:{url_args.pop('password', '')}"
                    auth = base64.b64encode(creds.encode('utf-8')).decode('utf-8')

                    # Create the auth request header, and add to url_args["headers"]
                    auth_header = {'Authorization': f'Basic {auth}'}
                    url_args["headers"] = {
                        **url_args.get("headers", {}),
                        **auth_header,
                    }

                    # extract any SSL-related args
                    urlopen_args = {
                        k: url_args.pop(k, None)
                        for k in "cafile capath context".split()
                    }

                data_request = urllib.request.Request(url=seqobj, **url_args)
                self._closeobj = urllib.request.urlopen(data_request, **urlopen_args)
                self._iterobj = _decoder(self._closeobj)
                self.type = ImportSourceType.url
            else:
                seqobj_path = Path(seqobj)
                if seqobj_path.suffixes[-2:] == [".tar", ".gz"]:
                    import tarfile

                    self._closeobj = tf = tarfile.open(seqobj, "r:gz")
                    inner_name = seqobj.removesuffix(".tar.gz")
                    try:
                        iterobj = tf.extractfile(inner_name)
                    except KeyError:
                        members = tf.getmembers()
                        if len(members) == 1:
                            iterobj = tf.extractfile(members[0])
                        else:
                            raise ValueError(
                                f"compressed tar archive contains multiple files, none matching {inner_name}"
                            )
                    self._iterobj = _decoder(iterobj)
                    self.type = ImportSourceType.tar_gzip
                elif seqobj_path.suffix == ".gz":
                    import gzip

                    self._closeobj = gzip.GzipFile(filename=seqobj)
                    self._iterobj = _decoder(self._closeobj)
                    self.type = ImportSourceType.gzip
                elif seqobj_path.suffix in (".xz", ".lzma"):
                    import lzma

                    self._iterobj = lzma.open(seqobj, "rt", encoding=encoding)
                    self.type = ImportSourceType.lzma
                elif seqobj_path.suffix == ".zip":
                    import zipfile

                    # assume file name inside zip is the same as the zip file without the trailing ".zip"
                    inner_name = Path(seqobj).stem
                    zipobj = zipfile.ZipFile(seqobj)
                    try:
                        self._closeobj = zipobj.open(inner_name)
                    except KeyError:
                        # if there is only one file in this zip file, use that
                        zip_contents = zipobj.infolist()
                        if len(zip_contents) == 1:
                            self._closeobj = zipobj.open(zip_contents[0].filename)
                        else:
                            raise ValueError(
                                f"zip archive contains multiple files, none matching {inner_name}"
                            )

                    self._iterobj = _decoder(self._closeobj)
                    self.type = ImportSourceType.zip

                elif seqobj_path.suffix in (".xlsx", ".xlsm", ".xlst"):
                    self._iterobj = open(seqobj, 'rb')
                    if self.type is None:
                        self.type = ImportSourceType.file
                else:
                    self._iterobj = open(seqobj, encoding=encoding)
                    if self.type is None:
                        self.type = ImportSourceType.file
        else:
            self._iterobj = iter(seqobj)
            self.type = ImportSourceType.iterable

    def __iter__(self) -> Iterator[Any]:
        return self

    def __next__(self) -> Any:
        return next(self._iterobj)

    def close(self) -> None:
        if hasattr(self._iterobj, "close"):
            self._iterobj.close()
        if self._closeobj is not None:
            self._closeobj.close()


FixedWidthParseSpec = Union[
    tuple[str, int],
    tuple[str, int, Optional[int]],
    tuple[str, int, Optional[int], Optional[Callable[[str], Any]]],
]


class FixedWidthReader:
    """
    Helper class to read fixed-width data and yield a sequence of dicts
    representing each row of data.

    Parameters:
    - slice_spec: a list of tuples defining each column in the input file
        Each tuple consists of:
        - data column label
        - starting column number (0-based)
        - (optional) ending column number; if omitted or None, uses the
          starting column of the next spec tuple as the ending column
          for this column
        - (optional) transform function: function called with the column
          string value to convert to other type, such as int, float,
          datetime, etc.; if omitted or None, str.strip() will be used
    - src_file: a string filename or a file-like object containing the
        fixed-width data to be loaded
    """

    def __init__(
        self,
        slice_spec: list[FixedWidthParseSpec],
        src_file: Union[str, Iterable, TextIO],
        encoding: str = "utf-8",
    ):
        def parse_spec(
            spec: list[FixedWidthParseSpec],
        ) -> list[tuple[str, slice, Callable[[str], Any]]]:
            def normalize_parse_spec(ps: FixedWidthParseSpec) -> FixedWidthParseSpec:
                return (*ps, None, None)[:4]  # noqa

            # add a "rest of the line" spec to the parse spec to terminate
            # the last column
            rest_of_line_spec = ("", None)
            spec.append(rest_of_line_spec)

            ret: list[tuple[str, slice, Callable[[str], Any]]] = []
            for cur, next_ in zip(spec, spec[1:]):
                label, col, end_col, fn = normalize_parse_spec(cur)
                if label is None:
                    continue
                if end_col is None:
                    end_col = next_[1]
                if fn is None:
                    fn = str.strip
                ret.append((label.lower(), slice(col, end_col), fn))
            return ret

        self._slices = parse_spec(slice_spec)
        self._src_file = src_file
        self._encoding = encoding

    def __iter__(self) -> Iterable[dict[str, Any]]:
        with contextlib.closing(_MultiIterator(self._src_file, self._encoding)) as _srciter:
            for line in _srciter:
                if not line.strip():
                    continue
                yield {label: fn(line[slc]) for label, slc, fn in self._slices}


def _make_comparator(cmp_fn: Callable[[Any, Any], bool]) -> Callable[[Any], Callable[[Any], Callable[[Any], bool]]]:
    """
    Internal function to help define Table.le, Table.lt, etc.
    """

    def comparator_with_value(value: Any) -> Callable[[Any], Callable[[Any], bool]]:
        def _table_comparator_fn(attr: str) -> Callable[[Any], bool]:
            def _inner(table_rec: Any) -> bool:
                try:
                    return cmp_fn(getattr(table_rec, attr), value)
                except TypeError:
                    return False

            _inner.__name__ = f"{cmp_fn.__name__}({attr}, {value!r})"
            return _inner

        _table_comparator_fn.fn = cmp_fn
        _table_comparator_fn.value = value
        _table_comparator_fn.is_comparator = True
        return _table_comparator_fn

    return comparator_with_value


def _make_comparator_none(cmp_fn: Callable[[Any, Any], bool]) -> Callable[[], Callable[[str], Callable[[Any], bool]]]:
    """
    Internal function to help define Table.is_none and Table.is_not_none.
    """

    def comparator_with_value() -> Callable[[str], Callable[[Any], bool]]:
        def _table_comparator_fn(attr: str) -> Callable[[Any], bool]:
            return lambda table_rec: cmp_fn(getattr(table_rec, attr), None)

        _table_comparator_fn.fn = cmp_fn
        _table_comparator_fn.value = None
        _table_comparator_fn.is_comparator = True
        return _table_comparator_fn

    return comparator_with_value


def _make_comparator_null(is_null: bool) -> Callable[[], Callable[[str], Callable[[Any], bool]]]:
    """
    Internal function to help define Table.is_null and Table.is_not_null.
    """

    def is_null_fn(a, value):
        return (a in (None, "")) == value

    def comparator_with_value() -> Callable[[str], Callable[[Any], bool]]:
        def _table_comparator_fn(attr: str) -> Callable[[Any], bool]:
            return lambda table_rec: is_null_fn(getattr(table_rec, attr, None), is_null)

        _table_comparator_fn.fn = is_null_fn
        _table_comparator_fn.value = is_null
        _table_comparator_fn.is_comparator = True
        _table_comparator_fn.__name__ = f"Table.is_null({is_null})"
        return _table_comparator_fn

    return comparator_with_value


def _make_comparator2(
        cmp_fn: Callable[[Any, Any, Any], bool]
) -> Callable[[Any, Any], Callable[[str], Callable[[Any], bool]]]:
    """
    Internal function to help define Table.within and between
    """

    def comparator_with_value(lower: Any, upper: Any) -> Callable[[str], Callable[[Any], bool]]:
        def _table_comparator_fn(attr: str) -> Callable[[Any], bool]:
            def _inner(table_rec: Any) -> bool:
                try:
                    return cmp_fn(lower, upper, getattr(table_rec, attr))
                except TypeError:
                    return False
            return _inner

        _table_comparator_fn.fn = cmp_fn
        _table_comparator_fn.lower = lower
        _table_comparator_fn.upper = upper
        _table_comparator_fn.is_comparator = True
        return _table_comparator_fn

    return comparator_with_value


def _make_comparator_regex(*reg_expr_args, **reg_expr_flags) -> Callable[[str], Callable[[Any], bool]]:
    warnings.warn(
        DeprecationWarning("Table.re_match(patt) comparator is deprecated,"
                           " replace with re.compile(patt).match"),
        stacklevel=2,
    )
    regex = re.compile(*reg_expr_args, **reg_expr_flags)
    cmp_fn = regex.match

    def _table_comparator_fn(attr: str) -> Callable[[Any], bool]:
        def _inner(table_rec):
            try:
                return cmp_fn(str(getattr(table_rec, attr, "")))
            except TypeError:
                return False
        return _inner

    _table_comparator_fn.fn = cmp_fn
    _table_comparator_fn.is_comparator = True
    return _table_comparator_fn


def _determine_suppressed_attrs(
        group_attrs: list[str], prev: tuple[Any, ...], curr: tuple[Any, ...],
        _compare=lambda attr_prev_curr: attr_prev_curr[1] == attr_prev_curr[2]
) -> set[str]:
    return {a for a, _, _ in itertools.takewhile(_compare, zip(group_attrs, prev, curr))}


TableContent = TypeVar("TableContent")


class Table(Generic[TableContent]):
    """
    Table is the main class in C{littletable}, for representing a collection of SimpleNamespaces or
    user-defined objects with publicly accessible attributes or properties.  Tables can be:
     - created, with an optional name, using standard Python L{C{Table() constructor}<__init__>}
     - indexed, with multiple indexes, with unique or non-unique values, see L{create_index}
     - queried, specifying values to exact match in the desired records, see L{where}
     - searched, specifying words to search for in values, where words may be found in text
       attributes, see L{create_search_index}
     - filtered (using L{where}), using a simple predicate function to match desired records;
       useful for selecting using inequalities or compound conditions
     - accessed directly for keyed values, using C{table.indexattribute[key]} - see L{__getattr__}
     - joined, using L{join_on} to identify attribute to be used for joining with another table, and
       L{join} or operator '+' to perform the actual join
     - pivoted, using L{pivot} to create a nested structure of sub-tables grouping objects
       by attribute values
     - grouped, using L{groupby} to create a summary table of computed values, grouped by a key
       attribute
     - L{imported<csv_import>}/L{exported<csv_export>} to CSV-format files; also supports working
       with TSV files, JSON files, and Excel spreadsheet files
    Queries and joins return their results as new Table objects, so that queries and joins can
    be easily performed as a succession of operations.
    """

    lt = staticmethod(_make_comparator(operator.lt))
    le = staticmethod(_make_comparator(operator.le))
    gt = staticmethod(_make_comparator(operator.gt))
    ge = staticmethod(_make_comparator(operator.ge))
    ne = staticmethod(_make_comparator(operator.ne))
    eq = staticmethod(_make_comparator(operator.eq))
    is_none = staticmethod(_make_comparator_none(operator.is_))
    is_not_none = staticmethod(_make_comparator_none(operator.is_not))
    is_null = staticmethod(_make_comparator_null(True))
    is_not_null = staticmethod(_make_comparator_null(False))
    is_in = staticmethod(_make_comparator(lambda x, seq: x in seq))
    not_in = staticmethod(_make_comparator(lambda x, seq: x not in seq))
    startswith = staticmethod(_make_comparator(lambda x, s: x is not None and str(x).startswith(s)))
    endswith = staticmethod(_make_comparator(lambda x, s: x is not None and str(x).endswith(s)))
    re_match = staticmethod(_make_comparator_regex)
    between = staticmethod(_make_comparator2(lambda lower, upper, x: x is not None and lower < x < upper))
    within = staticmethod(
        _make_comparator2(lambda lower, upper, x: x is not None and lower <= x <= upper)
    )
    in_range = staticmethod(
        _make_comparator2(lambda lower, upper, x: x is not None and lower <= x < upper)
    )

    INNER_JOIN = object()
    LEFT_OUTER_JOIN = object()
    RIGHT_OUTER_JOIN = object()
    FULL_OUTER_JOIN = object()
    _OUTER_JOIN_TYPES = (LEFT_OUTER_JOIN, RIGHT_OUTER_JOIN, FULL_OUTER_JOIN)

    @staticmethod
    def _wrap_dict(dd: Mapping[str, Any]) -> default_row_class:
        # do recursive wrap of dicts to namespace types
        ret = default_row_class(
            **{
                k: v if not isinstance(v, Mapping) else Table._wrap_dict(v)
                for k, v in dd.items()
            }
        )
        return ret

    @staticmethod
    def convert_numeric(
            s: Optional[str] = None,
            empty: Any = '',
            non_numeric: Type = object,
            force_float: bool = False,
            _int_fn: Callable[[str], int] = int,
    ) -> Union[Callable, Any]:
        """
        Convenience method for transforming columns of CSV data from str to float and/or int. By default,
        convert_numeric will convert int values to int, float values to float, and leave all other values as-is.

        Supported named args are:
        - empty - value to use for any value that is '' (such as "N/A" or "<missing>")
        - non_numeric - force a value for any value that is not int or float
        - force_float - convert all numerics to floats

        Examples::

            # default conversion behavior, can specify the function with or without ()'s
            tbl = lt.Table().csv_import(data, transforms={'value': lt.Table.convert_numeric})
            tbl = lt.Table().csv_import(data, transforms={'value': lt.Table.convert_numeric()})

            # force all non-numeric "value" entries to None
            tbl = lt.Table().csv_import(data, transforms={'value': lt.Table.convert_numeric(non_numeric=None)})

            # force all non-numeric "value" entries to 0
            tbl = lt.Table().csv_import(data, transforms={'value': lt.Table.convert_numeric(non_numeric=0)})

            # force all non-numeric or empty values to None
            tbl = lt.Table().csv_import(
                        data,
                        transforms={'value': lt.Table.convert_numeric(non_numeric=None, empty=None)}
                    )

            # force empty values to None (default is "")
            tbl = lt.Table().csv_import(data, transforms={'value': lt.Table.convert_numeric(empty=None)})

            # convert all numerics to float, even if ints
            tbl = lt.Table().csv_import(data, transforms={'value': lt.Table.convert_numeric(force_float=True)})

        """
        if s is None:
            # being "called" as part of calling csv_import, not actually converting anything here
            if empty != '' or non_numeric is not object or force_float:
                return partial(Table.convert_numeric,
                               empty=empty,
                               non_numeric=non_numeric,
                               force_float=force_float,
                               _int_fn=(int, float)[force_float])
            else:
                return Table.convert_numeric

        if s == '':
            return empty

        try:
            return _int_fn(s)
        except ValueError:
            try:
                return float(s)
            except ValueError:
                return s if non_numeric is object else non_numeric

    @staticmethod
    def parse_datetime(
            time_format: str,
            empty: Any = '',
            on_error: Optional[Any] = None
    ) -> Callable[[str], datetime.datetime]:
        """Convenience method to convert string data to a datetime.datetime instance,
           given a parse string (following strptime format).

           Can be used for transforming data loaded from CSV data sets.
        """

        def _converter(s: str = "") -> Any:
            if s == '':
                return empty
            try:
                return datetime.datetime.strptime(s, time_format)
            except ValueError:
                return on_error
        return _converter

    @staticmethod
    def parse_date(
            time_format: str, empty: Any = '', on_error: Optional[Any] = None
    ) -> Callable[[str], datetime.date]:
        """Convenience method to convert string data to a datetime.date instance,
           given a parse string (following strptime format).

           Can be used for transforming data loaded from CSV data sets.
        """

        def _converter(s: Optional[str] = None) -> Any:
            if not s:
                return empty
            try:
                return datetime.datetime.strptime(s, time_format).date()
            except ValueError:
                return on_error
        return _converter

    @staticmethod
    def parse_timedelta(
            time_format: str,
            reference_time: datetime.datetime = datetime.datetime.strptime("0:00:00", "%H:%M:%S"),
            empty: Any = '',
            on_error: Optional[Any] = None
    ) -> Callable[[str], datetime.timedelta]:
        """Convenience method to convert string data to a datetime.timedelta instance,
           given a parse string (following strptime format), and optionally a
           reference datetime.datetime.

           Can be used for transforming data loaded from CSV data sets.
        """

        def _converter(s: Optional[str] = None) -> Any:
            if s == '':
                return empty
            try:
                ret = datetime.datetime.strptime(s, time_format)
                return ret - reference_time
            except ValueError:
                return on_error
        return _converter

    def __init__(self, table_name: str = ""):
        """
        Create a new, empty Table.
        @param table_name: name for Table
        @type table_name: string (optional)
        """
        self(table_name)
        self.obs: list[Any] = []
        self._indexes: dict[str, _ObjIndex] = {}
        self._uniqueIndexes: list[_UniqueObjIndex] = []
        self._search_indexes: dict[str, dict[str, list]] = {}

        self.import_source_type: Optional[ImportSourceType] = None
        self.import_source: Optional[str] = None

        self.import_time = None
        self.create_time = datetime.datetime.now().astimezone(datetime.timezone.utc)
        self.modify_time = self.create_time

    @property
    def all(self) -> _TableAttributeValueLister:
        """
        Use C{'all'} to access all the values of a particular table column as a sequence.
        This is useful if passing the values on to another function that works with sequences
        of values::

            sum(customers.by.zipcode["12345"].all.order_total)

        The follow-on attribute C{'unique'} can be added to return a list of values with
        duplicates suppressed::

            customer_zip_codes = customers.all.zipcode.unique

        C{'all'} and C{'unique'} return a generator of the attribute values.
        """
        return _TableAttributeValueLister(self)

    @property
    def by(self) -> _IndexAccessor:
        """
        Use C{'by'} to access the Table's records by index key, as in::

            employees.by.socsecnum["000-00-0000"]
            customers.by.zipcode["12345"]

        The behavior differs slightly for unique and non-unique indexes:
         - if the index is unique, then retrieving a matching object, will return just the object;
           if there is no matching object, C{KeyError} is raised (making a table with a unique
           index behave very much like a Python dict)
         - if the index is non-unique, then all matching objects will be returned in a new Table,
           just as if a regular query had been performed; if no objects match the key value, an empty
           Table is returned and no exception is raised.

        If there is no index defined for the given attribute, then C{AttributeError} is raised.
        """
        return _IndexAccessor(self)

    @property
    def search(self) -> _TableSearcher:
        """
        Use C{'search'} to find records matching given query.
        Query is a list of keywords that may be found in a document. Keywords may be prefixed with
        "+" or "-" to indicate desired inclusion or exclusion. '++' and '--' will indicate a
        mandatory inclusion or exclusion. All other keywords will be optional,
        and will count toward a search score. Matching records will be returned in a list of
        (score, record) tuples, in descending order by score.

        Score will be computed as:
        - has '+' keyword:            1000
        - has '-' keyword:           -1000
        - has optional keyword:        100

        Note: the search index must have been previously created using create_search_index().
        If records have been added or removed, or text attribute contents modified since the
        search index was created, search will raise the SearchIndexInconsistentError exception.

        Parameters:
        - query - list of search keywords, with optional leading '++', '--', '+', or '-' flags
        - limit (optional) - limit the number of records returned
        - min_score (optional, default=0) - return only records with the given score or higher
        - include_words (optional, default=False) - also return the search index words for
          each record

        Example:

            # get top 10 recipes that have bacon but not anchovies
            recipes.search.ingredients("++bacon --anchovies", limit=10)

        """
        return _TableSearcher(self)

    def __len__(self) -> int:
        """Return the number of objects in the Table."""
        return len(self.obs)

    def __iter__(self) -> Iterator[TableContent]:
        """Create an iterator over the objects in the Table."""
        return iter(self.obs)

    def __getitem__(self, i: Any) -> Union[Table, TableContent]:
        """Provides direct indexed/sliced access to the Table's underlying list of objects."""
        if isinstance(i, slice):
            ret = self.copy_template()
            ret.insert_many(self.obs[i])
            return ret
        else:
            return self.obs[i]

    def __delitem__(self, i: Union[int, slice]) -> None:
        if isinstance(i, int):
            delidxs = [i]
        elif isinstance(i, slice):
            obs_len = len(self.obs)
            delidxs = sorted(list(range(*i.indices(obs_len))), reverse=True)
        else:
            raise TypeError("Table index must be int or slice")

        for idx in delidxs:
            self.pop(idx)

    def pop(self, i: int = -1) -> TableContent:
        ret = self.obs.pop(i)

        # remove from indexes
        for attr, ind in self._indexes.items():
            ind.remove(ret)

        self._contents_changed()
        return ret

    def __bool__(self) -> bool:
        return bool(self.obs)

    def __reversed__(self) -> Iterable[TableContent]:
        return reversed(self.obs)

    def __contains__(self, item: Union[TableContent, dict]) -> bool:
        if isinstance(item, Mapping):
            item = self._wrap_dict(item)
        return item in self.obs

    def index(self, item: Union[TableContent, dict]) -> int:
        if isinstance(item, dict):
            item = self._wrap_dict(item)
        return self.obs.index(item)

    def count(self, item: Union[TableContent, dict]) -> int:
        if isinstance(item, dict):
            item = self._wrap_dict(item)
        return self.obs.count(item)

    def __add__(self, other: Union[Table, _JoinTerm, Iterable]) -> Union[_JoinTerm, Table]:
        """Support UNION of 2 tables using "+" operator."""
        if isinstance(other, _JoinTerm):
            # special case if added to a JoinTerm, do join, not union
            return other + self
        elif isinstance(other, Table):
            # if other is another Table, just union them
            return self.union(other)
        else:
            # assume other is a sequence of some sort, insert all elements
            return self.clone().insert_many(other)

    def __iadd__(self, other: Table) -> Table:
        """Support UNION of 2 tables using "+=" operator."""
        return self.insert_many(other)

    def union(self, other: Table) -> Table:
        return self.clone().insert_many(other.obs)

    def __call__(self, table_name: Optional[str] = None) -> Table[TableContent]:
        """
        A simple way to assign a name to a table, such as those
        dynamically created by joins and queries.
        @param table_name: name for Table
        @type table_name: string
        """
        if table_name is not None:
            self.table_name = table_name
        return self

    def _attr_names(self) -> list[str]:
        return list(
            _object_attrnames(self.obs[0]) if self.obs else self._indexes.keys()
        )

    def copy_template(self, name: Optional[str] = None) -> Table[TableContent]:
        """
        Create empty copy of the current table, with copies of all
        index definitions.
        """
        ret: Table[TableContent] = Table(self.table_name)
        ret._indexes.update(
            {k: v.copy_template() for k, v in self._indexes.items()}
        )
        ret(name)
        return ret

    def clone(self, name: Optional[str] = None) -> Table[TableContent]:
        """
        Create full copy of the current table, including table contents
        and index definitions.
        """
        ret = self.copy_template().insert_many(self.obs)(name)
        return ret

    def create_index(
        self, attr: str, unique: bool = False, accept_none: bool = False, force: bool = False
    ) -> Table[TableContent]:
        """
        Create a new index on a given attribute.

        Having an index improves performance of sort and pivot methods.
        It also enables retrieving table contents using
        'table.by.<attr_name>[attr_value]' syntax.

        If this is a unique index, this makes the table act like a 'dict[K,T]',
        keyed by the values of the attr field. If not a unique index,
        then makes the table act like a 'defaultdict[Table[T]]', always returning
        a new Table of matching records (which may be empty if no records
        match).

        If C{unique} is True and records are found in the table with duplicate
        attribute values, the index is deleted and C{KeyError} is raised.

        If the table already has an index on the given attribute, then
        ValueError is raised.
        @param attr: the attribute to be used for indexed access and joins
        @type attr: string
        @param unique: flag indicating whether the indexed field values are
            expected to be unique across table entries
        @type unique: boolean
        @param accept_none: flag indicating whether None is an acceptable
            unique key value for this attribute (always True for non-unique
            indexes, default=False for unique indexes)
        @type accept_none: boolean
        @param force: flag indicating whether the index should be created
            even it if already exists (default = False)
        @type force: boolean
        """
        if force:
            self.drop_index(attr)

        if attr in self._indexes:
            raise ValueError(f"index {attr!r} already defined for table")

        if unique:
            self._indexes[attr] = _UniqueObjIndex(attr, accept_none)
            self._uniqueIndexes[:] = [
                ind for ind in self._indexes.values() if ind.is_unique
            ]
        else:
            self._indexes[attr] = _ObjIndex(attr)
            accept_none = True
        ind = self._indexes[attr]

        try:
            for obj in self.obs:
                obval = getattr(obj, attr, None)
                if obval is not None or accept_none:
                    ind[obval] = obj
                else:
                    raise KeyError("None is not an allowed key")
            return self

        except KeyError:
            self.drop_index(attr)
            raise

    def drop_index(self, attr: str) -> Table[TableContent]:
        """
        Deletes an index from the Table.  Can be used to drop and rebuild an index,
        or to convert a non-unique index to a unique index, or vice versa.
        @param attr: name of an indexed attribute
        @type attr: string
        """
        if attr in self._indexes:
            del self._indexes[attr]
            self._uniqueIndexes = [
                ind for ind in self._indexes.values() if ind.is_unique
            ]
        return self

    delete_index = drop_index

    def get_index(self, attr: str) -> _ReadonlyObjIndexWrapper:
        return _ReadonlyObjIndexWrapper(self._indexes[attr], self.copy_template())

    _NON_WORD_STRIPPER_RE = re.compile(r"[^\w_]?([\w._-]*)[^\w.]*")
    _NON_WORD_STRIPPER2_RE = re.compile(r"[^\w_-]?((?:\w|[-_]\w)+)(!>_-)$")
    _ACRONYM_WITH_PERIODS = re.compile(r"((?:\w\.){2,})(!>\.)$")
    _SIGNIFICANT_WORD_ENDING_RE = re.compile(rf"[a-z]{{2,}}({'|'.join(_significant_word_endings)})$")

    _PLURAL_ENDING_IN_IES = re.compile(r"(.*[^aeiouy])ies$")
    _PLURAL_ENDING_IN_ES = re.compile(r"(.*(?:ch|ss|sh|x))es$")
    _PLURAL_ENDING_IN_ES2 = re.compile(r"(.*(?:[bcdfghklmnprstuvwxz]|(qu))e)s$")
    _PLURAL_ENDING_IN_S = re.compile(r"(.*[^aeious])s$")
    _SINGULAR_ENDING_IN_S = re.compile(r"(.*(?:ness|ics))$")

    _PLURAL_MATCH_SUBS = (
        (_PLURAL_ENDING_IN_IES, r"\1y"),
        (_PLURAL_ENDING_IN_ES, r"\1"),
        (_PLURAL_ENDING_IN_ES2, r"\1"),
        (_SINGULAR_ENDING_IN_S, r"\1"),
        (_PLURAL_ENDING_IN_S, r"\1"),
    )

    @staticmethod
    def _normalize_word(s: str) -> str:
        match_res = [
            # an acronym of 2 or more "X." sequences, such as G.E. or I.B.M.
            re.compile(r"((?:\w\.){2,})"),
            # words that may be hyphenated or snake-case
            # (strip off any leading single non-word character)
            re.compile(r"[^\w_-]?((?:\w|[-_]\w)+)"),
        ]
        for match_re in match_res:
            if match := match_re.match(s):
                ret = match.group(1).lower().replace(".", "")
                return ret
        return ""

    @staticmethod
    def _normalize_word_gen(s: str, sw: frozenset) -> Iterable[str]:
        s = s.lower()
        if s in sw:
            return

        # strip non-word chars from front and back
        stripper = Table._NON_WORD_STRIPPER_RE
        s = stripper.match(s).group(1)

        if s in sw:
            return

        # catch plurals
        if (sa := s.rstrip(",.!?;:'\"-")).isalpha():
            s = sa
            if s in sw:
                return

            # check common plurals - if not found, check common plural patterns
            if not (sing := _plurals_map.get(s)):
                sing = next((sub_match[0] for re_sub, re_repl in Table._PLURAL_MATCH_SUBS
                            if (sub_match := re_sub.subn(re_repl, s))[1]),
                            s)

            if sing and sing != s:
                yield sing
            yield s

            # also add special ending words for code and documentation parsing
            if (
                s.endswith(_significant_word_endings)
                and (m := Table._SIGNIFICANT_WORD_ENDING_RE.match(s))
            ):
                yield m[1]

            return

        match_res = [
            # an acronym of 2 or more "X." sequences, such as G.E. or I.B.M.
            Table._ACRONYM_WITH_PERIODS,
            # words that may be hyphenated or snake-case
            # (strip off any leading single non-word character)
            Table._NON_WORD_STRIPPER2_RE,
        ]
        for match_re in match_res:
            if match := match_re.match(s):
                g1 = match.group(1).lower()
                yield g1.replace(".", "")
                for sep in "-":
                    if sep in g1:
                        yield from filter(None, g1.split(sep))
                break
        else:
            for sep in ".-":
                if sep in s:
                    yield from (
                        ss for ss in s.split(sep)
                        if len(ss) > 1
                    )
                    if sep == "." and all(len(ss) <= 1 for ss in s.split(".")):
                        yield s.replace(".", "")
            yield s

    @staticmethod
    def _normalize_split(s: str, sw: frozenset = frozenset()) -> Iterable[str]:
        return (
            ss for wd in s.split() for ss in Table._normalize_word_gen(wd, sw)
        )

    def create_search_index(
        self,
        attrname: str,
        *,
        using: Optional[str | Iterable[str]] = None,
        stopwords: Optional[Iterable[str]] = None,
        force: bool = False,
    ) -> Table[TableContent]:
        """
        Create a text search index for the given attribute.
        Regular indexes can perform range or equality checks against the
        value of a particular attribute. A search index will support
        calls to search() to perform keyword searches within the text
        of each record's attribute.

        Parameters:
        - attrname - name of attribute to be searched for word matches
        - stopwords (optional, default=defined list of English stop words)
          a user-defined list of stopwords to be filtered from
          search text and queries
        - force (optional, default=False) - force rebuild of a search index
          even if no records have been added or deleted (useful to rebuild
          a search index if a mutable record has been updated, which
          littletable cannot detect)

        Example:

            journal = Table()
            journal.insert(SimpleNamespace(date="1/1/2001", entry="The sky is so blue today."))
            journal.insert(SimpleNamespace(date="1/2/2001", entry="Feeling kind of blue."))
            journal.insert(SimpleNamespace(date="1/3/2001", entry="Delicious blueberry pie for dessert."))

            journal.create_search_index("entry")
            journal.search.entry("sky")
                [(namespace(date='1/1/2001', entry='The sky is so blue today.'), 100)]
            journal.search.entry("blue")
                [(namespace(date='1/1/2001', entry='The sky is so blue today.'), 100),
                 (namespace(date='1/2/2001', entry='Feeling kind of blue.'), 100)]
            journal.search.entry("blue --feeling")
                [(namespace(date='1/1/2001', entry='The sky is so blue today.'), 100)]
        """
        # if `using` argument supplied, add new field constructed from the names in `using`
        if using is not None:
            using_fields: list[str] = self._parse_fields_string(using)
            search_fields_getter = attrgetter(
                *using_fields,
                defaults=dict.fromkeys(using_fields, "")
            )
            search_field_builder = lambda r, _getter=search_fields_getter: " ".join(map(str, _getter(r))).strip()
            self.compute_field(attrname, search_field_builder)

        if attrname in self._search_indexes:
            if force or not self._search_indexes[attrname]["VALID"]:
                # stale search index, rebuild
                self._search_indexes.pop(attrname)
            else:
                return self

        stopwords_set: frozenset[str]
        if stopwords is None:
            stopwords_set = _stopwords
        else:
            stopwords_set = frozenset(stopwords)

        self._search_indexes[attrname] = defaultdict(list)
        new_index: dict[str, Any] = self._search_indexes[attrname]
        for i, obs_rec in enumerate(self.obs):
            if not (attrvalue := getattr(obs_rec, attrname, "")):
                continue
            words = self._normalize_split(attrvalue.lower(), stopwords_set)
            for wd in set(words):
                new_index[wd].append(i)

        # use uppercase keys for index metadata, since they should not
        # overlap with any search terms
        new_index["STOPWORDS"] = stopwords_set
        new_index["VALID"] = True

        return self

    def _search(
            self,
            attrname: str,
            query: str,
            limit: int = int(1e9),
            min_score: int = 0,
            include_words: bool = False,
            as_table: bool = True
    ):
        if attrname not in self._search_indexes:
            raise ValueError(f"no search index defined for attribute {attrname!r}")

        search_index = self._search_indexes[attrname]
        if not search_index["VALID"]:
            msg = (
                f"table has been modified since the search index for {attrname!r} was created,"
                " rebuild using create_search_index()"
            )
            raise SearchIndexInconsistentError(msg)
        stopwords = cast(frozenset[str], search_index["STOPWORDS"])

        plus_matches: dict[str, set[int]] = {}
        minus_matches: dict[str, set[int]] = {}
        opt_matches: dict[str, set[int]] = {}
        reqd_matches: set[int] = set()
        excl_matches: set[int] = set()
        reqd_words: dict[tuple[str, ...], dict[str, set[int]]] = {}

        if isinstance(query, str):
            query = shlex.split(query.strip())

        for keyword in query:
            keyword = keyword.lower()
            if keyword.startswith("++"):
                kwds = tuple(self._normalize_word_gen(keyword[2:], stopwords))
                reqd_words[kwds] = {}
                for kwd in kwds:
                    matched_entries = set(search_index.get(kwd, []))
                    reqd_words[kwds][kwd] = matched_entries
                    if not matched_entries:
                        continue

                    if kwd not in plus_matches:
                        plus_matches[kwd] = matched_entries

            elif keyword.startswith("--"):
                for kwd in self._normalize_word_gen(keyword[2:], stopwords):
                    excl_matches |= set(search_index.get(kwd, []))

            elif keyword.startswith("+"):
                for kwd in self._normalize_word_gen(keyword[1:], stopwords):
                    minus_matches.pop(kwd, None)
                    if kwd not in plus_matches and kwd not in reqd_matches:
                        plus_matches[kwd] = set(search_index.get(kwd, []))

            elif keyword.startswith("-"):
                for kwd in self._normalize_word_gen(keyword[1:], stopwords):
                    plus_matches.pop(kwd, None)
                    if kwd not in minus_matches and kwd not in excl_matches:
                        minus_matches[kwd] = set(search_index.get(kwd, []))

            else:
                for kwd in self._normalize_word_gen(keyword, stopwords):
                    if m := Table._SIGNIFICANT_WORD_ENDING_RE.match(keyword):
                        if kwd == m[1]:
                            continue
                    if kwd in plus_matches or kwd in minus_matches:
                        continue
                    opt_matches[kwd] = set(search_index.get(kwd, []))

        # process word groups to determine correct set of reqd_matches
        if reqd_words:
            reqd_matches = set(range(len(self.obs)))
            for reqd_word_tuple, word_matches_tuples in reqd_words.items():
                group_matches = set()
                for _, submatch in word_matches_tuples.items():
                    group_matches |= submatch
                if not group_matches:
                    # no possible match, force an impossible match set
                    reqd_matches = {-1}
                    break
                else:
                    reqd_matches &= group_matches

        # walk through plus, minus, and optional matches to build matching scores
        tally: Counter = Counter()
        for match_type, score in (
            (plus_matches, 1000),
            (minus_matches, -1000),
            (opt_matches, 100),
        ):
            for obj_set in match_type.values():
                if reqd_matches:
                    obj_set &= reqd_matches
                obj_set -= excl_matches
                for obj in obj_set:
                    tally[obj] += score

        # compose return structure, depending on whether the actual matched words in each entry should be included
        if include_words:
            ret = [
                (self[rec_idx], score,
                 sorted({}.fromkeys(self._normalize_split(getattr(self[rec_idx], attrname, ""))).keys() - stopwords))
                for rec_idx, score in tally.most_common(limit)
                if score > min_score]
        else:
            ret = [
                (self[rec_idx], score)
                for rec_idx, score in tally.most_common(limit)
                if score > min_score
            ]

        if as_table:
            tuple_ret = ret
            ret = self.copy_template()
            ret.table_name = " ".join(query)
            ret.insert_many(copy.copy(ret_rec[0]) for ret_rec in tuple_ret)
            score_attr = f"{attrname}_search_score"
            words_attr = f"{attrname}_search_words"
            try:
                for ret_rec, tup in zip(ret, tuple_ret):
                    setattr(ret_rec, score_attr, tup[1])
                    if len(tup) > 2:
                        setattr(ret_rec, words_attr, tup[2])
            except AttributeError:
                # not all record content types will accept new attributes
                pass

        return ret

    def delete_search_index(self, attrname: str) -> Table[TableContent]:
        """
        Deletes a previously-created search index on a particular attribute.
        """
        self._search_indexes.pop(attrname, None)
        return self

    def rebuild_search_index(self, attrname: str, force: bool = True) -> Table[TableContent]:
        """
        Rebuilds an existing search index if it has been invalidated, or if force=True.
        """
        try:
            existing_search_index = self._search_indexes[attrname]
        except KeyError:
            raise NoSuchIndexError(attrname)

        if not existing_search_index['VALID'] or force:
            self.create_search_index(
                attrname,
                stopwords=existing_search_index["STOPWORDS"],
                force=True,
            )

        return self

    def insert(self, obj: TableContent) -> Table[TableContent]:
        """
        Insert a new object into this Table.
        @param obj: any Python object -
        Objects can be constructed using any Python object; C{littletable}
        introspects the object's C{__dict__}, C{__slots__}, or C{_fields} attributes
        (or keys() result if pass a Python dict) to obtain join and
        index attributes and values.

        If a table is constructed using Python dicts, they are stored as
        C{types.SimpleNamespaces}, to support C{object.attribute} style access.

        If the table contains a unique index, and the record to be inserted would add
        a duplicate value for the indexed attribute, then C{KeyError} is raised, and the
        object is not inserted.

        If the table has no unique indexes, then it is possible to insert duplicate
        objects into the table.
        """
        return self.insert_many([obj])

    def insert_many(self, it: Iterable[TableContent]) -> Table[TableContent]:
        """Inserts a collection of objects into the table."""
        unique_indexes = self._uniqueIndexes
        NO_SUCH_ATTR = object()

        new_objs = it
        new_objs, first_obj = itertools.tee(new_objs)
        try:
            first = next(first_obj)
            if isinstance(first, dict):
                # passed in a list of dicts, save as attributed objects
                new_objs = (self._wrap_dict(obj) for obj in new_objs)
        except StopIteration:
            # iterator is empty, nothing to insert
            return self

        if unique_indexes:
            new_objs = list(new_objs)
            for ind in unique_indexes:
                ind_attr = ind.attr
                new_keys = {
                    getattr(obj, ind_attr, NO_SUCH_ATTR): obj for obj in new_objs
                }
                if not ind.accept_none and (
                    None in new_keys or NO_SUCH_ATTR in new_keys
                ):
                    raise KeyError(
                        f"unique key cannot be None or blank for index {ind_attr!r}",
                        [
                            ob
                            for ob in new_objs
                            if getattr(ob, ind_attr, NO_SUCH_ATTR) is None
                        ],
                    )
                if len(new_keys) < len(new_objs):
                    raise KeyError(
                        f"given sequence contains duplicate keys for index {ind_attr!r}"
                    )
                for key in new_keys:
                    if key in ind:
                        obj = new_keys[key]
                        raise KeyError(
                            f"duplicate unique key value {getattr(obj, ind_attr)!r} for index {ind_attr!r}",
                            new_keys[key],
                        )

        if self._indexes:
            for obj in new_objs:
                self.obs.append(obj)
                for attr, ind in self._indexes.items():
                    obval = getattr(obj, attr, None)
                    ind[obval] = obj
        else:
            self.obs.extend(new_objs)

        self._contents_changed()
        return self

    def remove(self, ob: Any) -> Table[TableContent]:
        """
        Removes an object from the table. If object is not in the table, then
        no action is taken and no exception is raised."""
        return self.remove_many([ob])

    def remove_many(self, it: Iterable) -> Table[TableContent]:
        """Removes a collection of objects from the table."""

        # if table is empty, there is nothing to remove
        if not self.obs:
            return self

        # find indicies of objects in iterable
        to_be_deleted = list(it)

        # if list of items to delete is empty, there is nothing to remove
        if not to_be_deleted:
            return self

        del_indices: list[int] = []
        for i, ob in enumerate(self.obs):
            if isinstance(ob, dict):
                ob = self._wrap_dict(ob)
            try:
                tbd_index = to_be_deleted.index(ob)
            except ValueError:
                continue
            else:
                del_indices.append(i)
                to_be_deleted.pop(tbd_index)

            # quit early if we have found them all
            if not to_be_deleted:
                break

        for i in sorted(del_indices, reverse=True):
            self.pop(i)

        if del_indices:
            self._contents_changed()

        return self

    def clear(self) -> Table[TableContent]:
        """
        Remove all contents from a Table and all indexes, but leave index definitions intact.
        """
        del self.obs[:]
        for idx in self._indexes.values():
            idx._clear()

        self._search_indexes.clear()
        return self

    def _contents_changed(self, *, invalidate_search_indexes: bool = True):
        """
        Internal method to be called whenever the contents of a table are modified.
        """
        if invalidate_search_indexes:
            for idx in self._search_indexes.values():
                idx["VALID"] = False  # noqa

        self.modify_time = datetime.datetime.now().astimezone(datetime.timezone.utc)

    def _query_attr_sort_fn(self, attr_val: tuple[str, Any]) -> int:
        """Used to order where keys by most selective key first"""
        attr, v = attr_val
        if attr in self._indexes:
            idx = self._indexes[attr]
            if v in idx:
                return len(idx[v])
            else:
                return 0
        else:
            return sys.maxsize

    def where(self, wherefn: Optional[PredicateFunction] = None, **kwargs: Any) -> Table[TableContent]:
        """
        Retrieves matching objects from the table, based on given
        named parameters.  If multiple named parameters are given, then
        only objects that satisfy all of the query criteria will be returned.

        @param wherefn: a method or lambda that returns a boolean result, as in::

           lambda ob : ob.unitprice > 10

        @type wherefn: callable(object) returning boolean

        @param kwargs: attributes for selecting records, given as additional
          named arguments of the form C{attrname="value"}, or C{Table.le(value)}
          using any of the methods C{Table.le}, C{Table.lt}, C{Table.ge}, C{Table.gt},
          C{Table.ne}, or C{Table.eq}, corresponding to the same methods defined
          in the stdlib C{operator} module.

        @return: a new Table containing the matching objects
        """
        if kwargs:
            # order query criteria in ascending order of number of matching items
            # for each individual given attribute; this will minimize the number
            # of filtering records that each subsequent attribute will have to
            # handle
            kwargs_list: list[tuple[str, Any]] = list(kwargs.items())
            if len(kwargs_list) > 1 and len(self) > 100:
                kwargs_list.sort(key=self._query_attr_sort_fn)

            ret = self
            NO_SUCH_ATTR = object()
            for k, v in kwargs_list:
                if callable(v):
                    if getattr(v, "is_comparator", False):
                        wherefn_k = v(k)
                    else:
                        def wherefn_k(obj):
                            try:
                                return v(getattr(obj, k, None))
                            except Exception:  # noqa
                                return False
                    newret = ret.where(wherefn_k)
                else:
                    newret = ret.copy_template()
                    if k in ret._indexes:
                        newret.insert_many(ret._indexes[k][v])
                    else:
                        newret.insert_many(
                            r for r in ret.obs if getattr(r, k, NO_SUCH_ATTR) == v
                        )

                ret = newret
                if not ret:
                    break
        else:
            ret = self

        if ret and wherefn is not None:
            newret = ret.copy_template()
            newret.insert_many(filter(wherefn, ret.obs))
            ret = newret

        if ret is self:
            ret = self.clone()

        return ret

    def delete(self, **kwargs: Any) -> int:
        """
        Deletes matching objects from the table, based on given
        named parameters.  If multiple named parameters are given, then
        only objects that satisfy all of the query criteria will be removed.
        @param kwargs: attributes for selecting records, given as additional
           named arguments of the form C{attrname="attrvalue"}, or
           C{attrname=Table.lt(value)} (see doc for L{Table.where}).
        @return: the number of objects removed from the table
        """
        if not kwargs:
            return 0

        affected = self.where(**kwargs)
        self.remove_many(affected)
        return len(affected)

    def shuffle(self) -> Table[TableContent]:
        """
        In-place random shuffle of the records in the table.
        """
        random.shuffle(self.obs)
        self._contents_changed()
        return self

    def orderby(
        self,
        key: Union[str, Iterable[str], Callable[[Any], Any]],
        reverse: bool = False,
    ) -> Table[TableContent]:
        """
        Sort Table in place, using given fields as sort key.
        @param key: if this is a string, it is a comma-separated list of field names,
           optionally followed by 'desc' to indicate descending sort instead of the
           default ascending sort; if a list or tuple, it is a list or tuple of field names
           or field names with ' desc' appended; if it is a function, then it is the
           function to be used as the sort key function
        @param reverse: (default=False) set to True if results should be in reverse order
        @type reverse: bool
        @return: self
        """
        if isinstance(key, (str, list, tuple)):
            attr_orders: list[tuple[str, str]]
            if isinstance(key, str):
                attrdefs = [s.strip() for s in key.split(",")]
                attr_orders = [(*a.split(), "asc")[:2] for a in attrdefs]  # noqa
            else:
                # attr definitions were already resolved to a sequence by the caller
                if isinstance(key[0], str):
                    attr_orders = [(*a.split(), "asc")[:2] for a in key]  # noqa
                else:
                    attr_orders = key
            attrs = [attr for attr, order in attr_orders]

            # special optimization if all orders are ascending or descending
            if all(order == "asc" for attr, order in attr_orders):
                for seq in (self.obs, *self._indexes.values()):
                    seq.sort(key=attrgetter(*attrs), reverse=reverse)
            elif all(order == "desc" for attr, order in attr_orders):
                for seq in (self.obs, *self._indexes.values()):
                    seq.sort(key=attrgetter(*attrs), reverse=not reverse)
            else:
                # mix of ascending and descending sorts, have to do succession of sorts
                # leftmost attr is the most primary sort key, so reverse attr_orders to do
                # succession of sorts from right to left
                for seq in (self.obs, *self._indexes.values()):
                    for attr, order in reversed(attr_orders):
                        seq.sort(
                            key=attrgetter(attr), reverse=(order == "desc")
                        )
        else:
            # sorting given a sort key function
            keyfn = key
            self.obs.sort(key=keyfn, reverse=reverse)

        self._contents_changed()
        return self

    # backward-compatibility name
    sort = orderby

    def rank(self, rank_col_name: str = "rank", start=1) -> Table:
        """
        Add ranking column to each row in the table.

        :param rank_col_name: name to give the new column; default="rank"
        :param start: initial ranking number; default=1
        """
        for i, ob in enumerate(self.obs, start=start):
            setattr(ob, rank_col_name, i)
        return self

    def select(
            self,
            fields: Optional[Union[Iterable[str], str]] = None,
            **exprs: Callable[[TableContent], Any]
    ) -> Table:
        """
        Create a new table containing a subset of attributes, with optionally
        newly-added fields computed from each rec in the original table.

        @param fields: list of strings, or single space-delimited string, listing attribute name to be included in the
        output
         - names starting with '-' indicate to suppress that field
         - '*' means include all other field names
         - if no fields are specifically included, then all fields are used
        @type fields: list, or space-delimited string
        @param exprs: one or more named callable arguments, to compute additional fields using the given function
        @type exprs: C{name=callable}, callable takes the record as an argument, and returns the new attribute value
        If a string is passed as a callable, this string will be used using string formatting, given the record
        as a source of interpolation values.  For instance, C{fullName = '%(lastName)s, %(firstName)s'}

        """
        if fields is not None:
            fields = self._parse_fields_string(fields)
        else:
            fields = []

        def _make_string_callable(expr):
            if isinstance(expr, str):
                return (
                    lambda r: expr.format(r)
                    if not isinstance(r, (list, tuple))
                    else expr.format(*r)
                )
            else:
                return expr

        exprs = {k: _make_string_callable(v) for k, v in exprs.items()}

        raw_tuples = []
        attrvalues_getter = attrgetter(*fields) if fields else lambda _: ()
        for ob in self.obs:
            attrvalues = attrvalues_getter(ob)
            if exprs:
                attrvalues += tuple(expr(ob) for expr in exprs.values())
            raw_tuples.append(attrvalues)

        all_names = (*fields, *exprs)
        ret: Table[TableContent] = Table(self.table_name)
        ret._indexes.update(
            {k: v.copy_template() for k, v in self._indexes.items() if k in all_names}
        )
        if self:
            ret.insert_many(
                default_row_class(**dict(zip(all_names, out_tuple)))
                for out_tuple in raw_tuples
            )
        return ret

    def formatted_table(self, *fields: str, **exprs: str) -> Table:
        """
        Create a new table with all string formatted attribute values, typically in preparation for
        formatted output.
        @param fields: one or more strings, each string is an attribute name to be included in the output
        @type fields: string (multiple)
        @param exprs: one or more named string arguments, to format the given attribute with a formatting string
        @type exprs: name=string
        """
        select_exprs = dict()
        for fld in fields:
            if fld not in select_exprs:
                select_exprs[fld] = lambda r, f=fld: str(getattr(r, f, "None"))

        for ename, expr in exprs.items():
            if isinstance(expr, str):
                if re.match(r"[a-zA-Z_][a-zA-Z0-9_]*$", expr):
                    select_exprs[ename] = lambda r, expr_=expr: str(getattr(r, expr_, "None"))
                else:
                    if "{}" in expr or "{0}" or "{0:" in expr:
                        select_exprs[ename] = lambda r, expr_=expr: expr_.format(r)
                    else:
                        select_exprs[ename] = lambda r, expr_=expr: expr_.format(
                            getattr(r, ename, "None")
                        )

        return self.select(**select_exprs)

    def format(self, fmt: str):
        """
        Generates a list of strings, one for each row in the table, using the input string
        as a format template for printing out a single row.
        """
        for line in self:
            yield fmt.format(**_to_dict(line))

    def join(
        self,
        other,
        attrlist: Optional[Union[str, Iterable[str]]] = None,
        auto_create_indexes: bool = True,
        **kwargs: Any,
    ) -> Table:
        """
        Join the objects of one table with the objects of another, based on the given
        matching attributes in the named arguments.  The attrlist specifies the attributes to
        be copied from the source tables - if omitted, all attributes will be copied.  Entries
        in the attrlist may be single attribute names, or if there are duplicate names in both
        tables, then a C{(table,attributename)} tuple can be given to disambiguate which
        attribute is desired. A C{(table,attributename,alias)} tuple can also be passed, to
        rename an attribute from a source table.

        This method may be called directly, or can be constructed using the L{join_on} method and
        the '+' operator.  Using this syntax, the join is specified using C{table.join_on("xyz")}
        to create a JoinTerm containing both table and joining attribute.  Multiple JoinTerm
        or tables can be added to construct a compound join expression.  When complete, the
        join expression gets executed by calling the resulting join definition,
        using C{join_expression([attrlist])}.

        @param other: other table to join to
        @param attrlist: list of attributes to be copied to the new joined table; if
            none provided, all attributes of both tables will be used (taken from the first
            object in each table)
        @type attrlist: string, or list of strings or C{(table,attribute[,alias])} tuples
            (list may contain both strings and tuples)
        @param kwargs: attributes to join on, given as additional named arguments
            of the form C{table1attr="table2attr"}, or a dict mapping attribute names.
        @type auto_create_indexes: bool
        @param auto_create_indexes: flag to simplify joining tables, to automatically
            create necessary indexes instead of raising ValueError if a join field
            is not yet indexed (default=True)
        @returns: a new Table containing the joined data as new SimpleNamespaces
        """
        if not kwargs:
            raise TypeError(
                "must specify at least one join attribute as a named argument"
            )
        this_cols, other_cols = list(kwargs.keys()), list(kwargs.values())

        if not all(isinstance(col, str) for col in this_cols) or not all(
            isinstance(col, str) for col in other_cols
        ):
            raise TypeError("all join keywords must be of type str")

        retname = (
            f"({self.table_name}:{'/'.join(this_cols)}"
            "^"
            f"{other.table_name}:{'/'.join(other_cols)})"
        )

        # if inner join, make sure both tables contain records to join - if not, just return empty list
        if not (self.obs and other.obs):
            return Table(retname)

        attr_spec_list = attrlist
        if isinstance(attrlist, str):
            attr_spec_list = re.split(r"[,\s]+", attrlist)

        # expand attrlist to full (table, name, alias) tuples
        full_attr_specs: list[tuple[Table, str, str]]
        if attr_spec_list is None:
            full_attr_specs = [(self, namestr, namestr) for namestr in self._attr_names()]
            full_attr_specs += [(other, namestr, namestr) for namestr in other._attr_names()]
        else:
            full_attr_specs = []
            this_attr_names = set(self._attr_names())
            other_attr_names = set(other._attr_names())
            for attr_spec in attr_spec_list:
                if isinstance(attr_spec, tuple):
                    # assume attr_spec contains at least (table, col_name), fill in alias if missing
                    # to be same as col_name
                    if len(attr_spec) == 2:
                        attr_spec = (*attr_spec, attr_spec[1])
                    full_attr_specs.append(attr_spec)  # noqa
                else:
                    name = attr_spec
                    if name in this_attr_names:
                        full_attr_specs.append((self, name, name))
                    elif attr_spec in other_attr_names:
                        full_attr_specs.append((other, name, name))
                    else:
                        raise ValueError(f"join attribute not found: {name!r}")

        # regroup attribute specs by table
        this_attr_specs = [
            attr_spec for attr_spec in full_attr_specs if attr_spec[0] is self
        ]
        other_attr_specs = [
            attr_spec for attr_spec in full_attr_specs if attr_spec[0] is other
        ]

        if auto_create_indexes:
            for tbl, col_list in ((self, this_cols), (other, other_cols)):
                for col in col_list:
                    if col not in tbl._indexes:
                        tbl.create_index(col)
        else:
            # make sure all join columns are indexed
            unindexed_cols: list[str] = []
            for tbl, col_list in ((self, this_cols), (other, other_cols)):
                unindexed_cols.extend(
                    col for col in col_list if col not in tbl._indexes
                )
            if unindexed_cols:
                raise ValueError(
                    f"indexed attributes required for join: {','.join(unindexed_cols)}"
                )

        # find matching rows
        matching_rows: list[tuple[Table, Table]] = []
        key_map_values = list(
            zip(this_cols, other_cols, (self._indexes[key].keys() for key in this_cols))
        )
        for join_values in itertools.product(*(kmv[-1] for kmv in key_map_values)):
            base_this_where_dict = dict(zip(this_cols, join_values))
            base_other_where_dict = dict(zip(other_cols, join_values))

            # compute inner join rows to start
            this_rows = self.where(**base_this_where_dict)
            other_rows = other.where(**base_other_where_dict)

            matching_rows.append((this_rows, other_rows))

        # remove attr_specs from other_attr_specs if alias is duplicate of any alias in this_attr_specs
        this_attr_specs_aliases = {alias for tbl, col, alias in this_attr_specs}
        other_attr_specs = [
            (tbl, col, alias)
            for tbl, col, alias in other_attr_specs
            if alias not in this_attr_specs_aliases
        ]

        join_rows: list[Any] = []
        for this_rows, other_rows in matching_rows:
            for trow, orow in itertools.product(this_rows, other_rows):
                retobj = default_row_class()
                for _, attr_name, alias in this_attr_specs:
                    setattr(retobj, alias, getattr(trow, attr_name, None))
                for _, attr_name, alias in other_attr_specs:
                    setattr(retobj, alias, getattr(orow, attr_name, None))
                join_rows.append(retobj)

        ret: Table[TableContent] = Table(retname)
        ret.insert_many(join_rows)

        # add indexes as defined in source tables
        for tbl, attr_name, alias in this_attr_specs + other_attr_specs:
            if attr_name in tbl._indexes:
                if alias not in ret._indexes:
                    ret.create_index(alias)  # no unique indexes in join results

        return ret

    def outer_join(
        self,
        join_type,
        other: Table,
        attrlist: Optional[Union[Iterable[str], str]] = None,
        auto_create_indexes: bool = True,
        **kwargs: Any,
    ):
        """
        Join the objects of one table with the objects of another, based on the given
        matching attributes in the named arguments.  The attrlist specifies the attributes to
        be copied from the source tables - if omitted, all attributes will be copied.  Entries
        in the attrlist may be single attribute names, or if there are duplicate names in both
        tables, then a C{(table,attributename)} tuple can be given to disambiguate which
        attribute is desired. A C{(table,attributename,alias)} tuple can also be passed, to
        rename an attribute from a source table.

        This method may be called directly, or can be constructed using the L{join_on} method and
        the '+' operator.  Using this syntax, the join is specified using C{table.join_on("xyz")}
        to create a JoinTerm containing both table and joining attribute.  Multiple JoinTerm
        or tables can be added to construct a compound join expression.  When complete, the
        join expression gets executed by calling the resulting join definition,
        using C{join_expression([attrlist])}.

        @param join_type: type of outer join to be performed
        @type join_type: must be Table.LEFT_OUTER_JOIN, Table.RIGHT_OUTER_JOIN, or Table.FULL_OUTER_JOIN
        @param other: other table to join to
        @param attrlist: list of attributes to be copied to the new joined table; if
            none provided, all attributes of both tables will be used (taken from the first
            object in each table)
        @type attrlist: string, or list of strings or C{(table,attribute[,alias])} tuples
            (list may contain both strings and tuples)
        @param kwargs: attributes to join on, given as additional named arguments
            of the form C{table1attr="table2attr"}, or a dict mapping attribute names.
        @returns: a new Table containing the joined data as new SimpleNamespaces
        """
        if join_type not in Table._OUTER_JOIN_TYPES:
            join_names = [
                nm
                for nm, join_var in vars(Table).items()
                if join_var in Table._OUTER_JOIN_TYPES
            ]
            raise ValueError(
                f"join argument must be one of [{', '.join(f'Table.{nm}' for nm in join_names)}]"
            )

        if not kwargs:
            raise TypeError(
                "must specify at least one join attribute as a named argument"
            )
        this_cols, other_cols = list(kwargs.keys()), list(kwargs.values())

        if not all(isinstance(col, str) for col in this_cols) or not all(
            isinstance(col, str) for col in other_cols
        ):
            raise TypeError("all join keywords must be of type str")

        retname = (
            f"({self.table_name}:{'/'.join(this_cols)}"
            "|"
            f"{other.table_name}:{'/'.join(other_cols)})"
        )

        if self is other:
            return self.clone()(retname)

        attr_spec_list = attrlist
        if isinstance(attrlist, str):
            attr_spec_list = re.split(r"[,\s]+", attrlist)

        # expand attrlist to full (table, name, alias) tuples
        full_attr_specs: list[tuple[Table, str, str]]
        if attr_spec_list is None:
            full_attr_specs = [(self, namestr, namestr) for namestr in self._attr_names()]
            full_attr_specs += [(other, namestr, namestr) for namestr in other._attr_names()]
        else:
            full_attr_specs = []
            this_attr_names = set(self._attr_names())
            other_attr_names = set(other._attr_names())
            for attr_spec in attr_spec_list:
                if isinstance(attr_spec, tuple):
                    # assume attr_spec contains at least (table, col_name), fill in alias if missing
                    # to be same as col_name
                    if len(attr_spec) == 2:
                        attr_spec = attr_spec + (attr_spec[-1],)
                    full_attr_specs.append(attr_spec)
                else:
                    name = attr_spec
                    if name in this_attr_names:
                        full_attr_specs.append((self, name, name))
                    elif attr_spec in other_attr_names:
                        full_attr_specs.append((other, name, name))
                    else:
                        raise ValueError(f"join attribute not found: {name!r}")

        # regroup attribute specs by table
        this_attr_specs = [
            attr_spec for attr_spec in full_attr_specs if attr_spec[0] is self
        ]
        other_attr_specs = [
            attr_spec for attr_spec in full_attr_specs if attr_spec[0] is other
        ]

        if auto_create_indexes:
            for tbl, col_list in ((self, this_cols), (other, other_cols)):
                for col in col_list:
                    if col not in tbl._indexes:
                        tbl.create_index(col)
        else:
            # make sure all join columns are indexed
            unindexed_cols: list[str] = []
            for tbl, col_list in ((self, this_cols), (other, other_cols)):
                unindexed_cols.extend(
                    col for col in col_list if col not in tbl._indexes
                )
            if unindexed_cols:
                raise ValueError(
                    f"indexed attributes required for join: {','.join(unindexed_cols)}"
                )

        # find matching rows
        matching_rows: list[tuple[Table, Table]] = []
        if join_type == Table.RIGHT_OUTER_JOIN:
            key_map_values = list(
                zip(
                    this_cols,
                    other_cols,
                    (self._indexes[key].keys() for key in this_cols),
                )
            )
        elif join_type == Table.LEFT_OUTER_JOIN:
            key_map_values = list(
                zip(
                    this_cols,
                    other_cols,
                    (other._indexes[key].keys() for key in other_cols),
                )
            )
        else:
            key_map_values = list(
                zip(
                    this_cols,
                    other_cols,
                    (
                        set(self._indexes[this_key].keys())
                        | set(other._indexes[other_key].keys())
                        for this_key, other_key in zip(this_cols, other_cols)
                    ),
                )
            )

        for join_values in itertools.product(*(kmv[-1] for kmv in key_map_values)):
            base_this_where_dict = dict(zip(this_cols, join_values))
            base_other_where_dict = dict(zip(other_cols, join_values))

            # compute inner join rows to start
            this_rows = self.where(**base_this_where_dict)
            other_rows = other.where(**base_other_where_dict)

            if join_type in (Table.FULL_OUTER_JOIN, Table.LEFT_OUTER_JOIN):
                if not this_rows:
                    this_outer_dict = dict.fromkeys(this_cols, None)
                    this_outer_dict.update(dict(zip(this_cols, join_values)))
                    this_rows.insert(default_row_class(**this_outer_dict))

            if join_type in (Table.FULL_OUTER_JOIN, Table.RIGHT_OUTER_JOIN):
                if not other_rows:
                    other_outer_dict = dict.fromkeys(other_cols, None)
                    other_outer_dict.update(dict(zip(other_cols, join_values)))
                    other_rows.insert(default_row_class(**other_outer_dict))

            matching_rows.append((this_rows, other_rows))

        # remove attr_specs from other_attr_specs if alias is duplicate of any alias in this_attr_specs
        this_attr_specs_aliases = {alias for tbl, col, alias in this_attr_specs}
        other_attr_specs = [
            (tbl, col, alias)
            for tbl, col, alias in other_attr_specs
            if alias not in this_attr_specs_aliases
        ]

        join_rows: list[Any] = []
        for this_rows, other_rows in matching_rows:
            for trow, orow in itertools.product(this_rows, other_rows):
                retobj = default_row_class()
                for _, attr_name, alias in this_attr_specs:
                    setattr(retobj, alias, getattr(trow, attr_name, None))
                for _, attr_name, alias in other_attr_specs:
                    setattr(retobj, alias, getattr(orow, attr_name, None))
                join_rows.append(retobj)

        ret: Table = Table(retname)
        ret.insert_many(join_rows)

        # add indexes as defined in source tables
        for tbl, attr_name, alias in this_attr_specs + other_attr_specs:
            if attr_name in tbl._indexes:
                if alias not in ret._indexes:
                    ret.create_index(alias)  # no unique indexes in join results

        return ret

    def join_on(self, attr: str, join: str = "inner") -> _JoinTerm:
        """
        Creates a JoinTerm in preparation for joining with another table, to
        indicate what attribute should be used in the join.  Only indexed attributes
        may be used in a join.
        @param attr: attribute name to join from this table (may be different
            from the attribute name in the table being joined to)
        @type attr: string
        @returns: L{JoinTerm}"""
        if attr not in self._indexes:
            raise ValueError("can only join on indexed attributes")
        return _JoinTerm(self, attr, join)

    def pivot(self, attrlist: Union[Iterable[str], str]) -> _PivotTable:
        """
        Pivots the data using the given attributes, returning a L{PivotTable}.
        @param attrlist: list of attributes to be used to construct the pivot table
        @type attrlist: list of strings, or string of space-delimited attribute names
        """
        if isinstance(attrlist, str):
            attrlist = attrlist.split()
        else:
            attrlist = list(attrlist)

        if all(a in self._indexes for a in attrlist):
            return _PivotTable(self, [], attrlist)
        else:
            missing = set(attrlist) - self._indexes.keys()
            raise ValueError(
                f"pivot can only be called using indexed attributes;"
                f" {', '.join(repr(m) for m in missing)} not indexed"
            )

    def _import(
        self,
        source: _ImportExportDataContainer,
        encoding: str = "utf-8",
        transforms: Optional[dict] = None,
        filters: Optional[dict] = None,
        reader=csv.DictReader,
        row_class: Optional[type] = None,
        limit: Optional[int] = None,
        url_args: Optional[dict] = None,
    ) -> Table:

        if row_class is None:
            row_class = default_row_class

        with contextlib.closing(_MultiIterator(source, encoding, url_args)) as _srciter:
            csvdata = reader(_srciter)

            if transforms:
                transformers: list[tuple[str, Callable[[Any], Any], Any]] = []
                for k, v in transforms.items():
                    if k == "*":
                        continue
                    if isinstance(v, tuple):
                        v, default = v
                    else:
                        default = None
                    if callable(v):
                        transformers.append((k, v, default))
                    else:
                        transformers.append((k, lambda __: v, default))

                def transformer(csv_rec, xformers=transformers):  # noqa
                    for xform_k, xform_v, xform_default in xformers:
                        try:
                            csv_rec[xform_k] = xform_v(csv_rec[xform_k])
                        except Exception:  # noqa
                            csv_rec[xform_k] = xform_default
                    return csv_rec

                no_default_given = object()
                non_star_transforms = transforms.keys() - {"*"}
                def catch_all_transformer(csv_rec, xform_fn, xform_default):
                    for csv_k, csv_v in csv_rec.items():
                        if csv_k in non_star_transforms:
                            continue
                        try:
                            csv_rec[csv_k] = xform_fn(csv_v)
                        except Exception:  # noqa
                            if xform_default is not no_default_given:
                                csv_rec[csv_k] = xform_default
                    return csv_rec

                csvdata = (transformer(d) for d in csvdata)

                if (wild_card := transforms.get("*")) is not None:
                    if isinstance(wild_card, tuple):
                        wild_card_fn, wild_card_default = wild_card
                    else:
                        wild_card_fn, wild_card_default = wild_card, no_default_given
                    csvdata = (
                        catch_all_transformer(d, wild_card_fn, wild_card_default)
                        for d in csvdata
                    )

            if filters:
                for k, v in filters.items():
                    if callable(v):
                        if getattr(v, "is_comparator", False):
                            # comparators work against attrs, but csvdata is still just a series of
                            # dicts, so must convert each to a temporary row_class instance to perform the
                            # comparator predicate method
                            fn = getattr(v, "fn")
                            no_object = object()
                            value = getattr(v, "value", no_object)
                            upper = getattr(v, "upper", no_object)
                            lower = getattr(v, "lower", no_object)
                            if value is not no_object:
                                csvdata = filter(
                                    lambda rec_dict: fn(rec_dict.get(k), value), csvdata
                                )
                            elif upper is not no_object:
                                filt_fn = partial(fn, lower, upper)
                                csvdata = filter(
                                    lambda rec_dict: (k_val := rec_dict.get(k)) is not None and filt_fn(k_val),
                                    csvdata,
                                )
                            else:
                                csvdata = filter(
                                    lambda rec_dict: fn(rec_dict.get(k)), csvdata
                                )

                        else:
                            csvdata = filter(lambda csv_rec: v(csv_rec.get(k)), csvdata)
                    else:
                        csvdata = filter(lambda csv_rec: csv_rec.get(k) == v, csvdata)

            if limit is not None:
                csvdata = itertools.islice(csvdata, limit)

            self.insert_many(row_class(**s) for s in csvdata)

            self.import_source_type = _srciter.type
            if self.import_source_type in (ImportSourceType.path,
                                           ImportSourceType.file,
                                           ImportSourceType.url,
                                           ImportSourceType.zip,
                                           ImportSourceType.gzip,
                                           ImportSourceType.tar_gzip,
                                           ImportSourceType.lzma,
                                           ):
                # if url, strip off query args if any
                if self.import_source_type is ImportSourceType.url:
                    source = str(source).partition("?")[0]
                self.import_source = str(source)
                self(str(source))

        self.import_time = datetime.datetime.now().astimezone(datetime.timezone.utc)
        self._contents_changed()
        return self

    def csv_import(
        self,
        csv_source: _ImportExportDataContainer,
        encoding: str = "utf-8",
        transforms: Optional[dict] = None,
        filters: Optional[dict] = None,
        row_class: Optional[type] = None,
        limit: Optional[int] = None,
        fieldnames: Optional[Union[Iterable[str], str]] = None,
        **kwargs: Any,
    ) -> Table:
        """
        Imports the contents of a CSV-formatted file into this table.
        @param csv_source: CSV file - if a string is given, the file with that name will be
            opened, read, and closed; if a file object is given, then that object
            will be read as-is, and left for the caller to be closed.
        @type csv_source: string or file
        @param encoding: encoding to be used for reading source text if C{csv_source} is
            passed as a string filename
        @type encoding: string (default='UTF-8')
        @param transforms: dict of functions by attribute name; if given, each
            attribute will be transformed using the corresponding transform; if there is no
            matching transform, the attribute will be read as a string (default); the
            transform function can also be defined as a (function, default-value) tuple; if
            there is an Exception raised by the transform function, then the attribute will
            be set to the given default value
        @type transforms: dict (optional)
        @param filters: dict of functions by attribute name; if given, each
            newly-read record will be filtered before being added to the table, with each
            filter function run using the corresponding attribute; if any filter function
            returns False, the record is not added to the table. Useful when reading large
            input files, to pre-screen only for data matching one or more filters
        @type filters: dict (optional)
        @param row_class: class to construct for each imported row when populating table (default=SimpleNamespace)
        @type row_class: type
        @param limit: number of records to import
        @type limit: int (optional)
        @param kwargs: additional constructor arguments for csv C{DictReader} objects, such as C{delimiter}
            or C{fieldnames}; these are passed directly through to the csv C{DictReader} constructor. kwargs
            may also contain any of the following if importing using a URL: C{headers}, C{data}, C{username},
            C{password}
        @type kwargs: named arguments (optional)
        @param fieldnames: names for imported columns; used if there is no header line in the input file
        @type fieldnames: list[str] or str
        """
        non_reader_args = (
            "encoding csv_source transforms row_class limit headers data username password cafile"
            " capath context".split()
        )
        url_arg_names = "headers data username password cafile capath context".split()
        url_args = {k: kwargs.pop(k) for k in url_arg_names if k in kwargs}
        reader_args = {
            k: v for k, v in kwargs.items() if k not in non_reader_args
        }
        reader_args["fieldnames"] = fieldnames.split() if isinstance(fieldnames, str) else fieldnames
        return self._import(
            csv_source,
            encoding=encoding,
            transforms=transforms,
            filters=filters,
            reader=lambda src: csv.DictReader(src, **reader_args),
            row_class=row_class,
            limit=limit,
            url_args=url_args,
        )

    def _xsv_import(
        self,
        xsv_source: _ImportExportDataContainer,
        encoding: str = "utf-8",
        transforms: Optional[dict] = None,
        filters: Optional[dict] = None,
        row_class: Optional[type] = None,
        limit: Optional[int] = None,
        **kwargs: Any,
    ) -> Table:
        non_reader_args = (
            "encoding xsv_source transforms row_class limit filters headers data username password"
            " cafile capath context".split()
        )
        url_arg_names = "headers data username password cafile capath context".split()
        url_args = {k: kwargs.pop(k) for k in url_arg_names if k in kwargs}
        reader_args = {
            k: v for k, v in kwargs.items() if k not in non_reader_args
        }
        return self._import(
            xsv_source,
            encoding=encoding,
            transforms=transforms,
            filters=filters,
            reader=lambda src: csv.DictReader(src, **reader_args),
            row_class=row_class,
            limit=limit,
            url_args=url_args,
        )

    def tsv_import(
        self,
        xsv_source: _ImportExportDataContainer,
        encoding: str = "utf-8",
        transforms: Optional[dict] = None,
        filters: Optional[dict] = None,
        row_class: Optional[type] = None,
        limit: Optional[int] = None,
        fieldnames: Optional[Union[Iterable[str], str]] = None,
        **kwargs,
    ) -> Table:
        """
        Imports the contents of a tab-separated data file into this table.
        @param xsv_source: tab-separated data file - if a string is given, the file with that name will be
            opened, read, and closed; if a file object is given, then that object
            will be read as-is, and left for the caller to be closed.
        @type xsv_source: string or file
        @param transforms: dict of functions by attribute name; if given, each
            attribute will be transformed using the corresponding transform; if there is no
            matching transform, the attribute will be read as a string (default); the
            transform function can also be defined as a (function, default-value) tuple; if
            there is an Exception raised by the transform function, then the attribute will
            be set to the given default value
        @type transforms: dict (optional)
        @param row_class: class to construct for each imported row when populating table (default=SimpleNamespace)
        @type row_class: type
        @param limit: number of records to import
        @type limit: int (optional)
        @param fieldnames: names for imported columns; used if there is no header line in the input file
        @type fieldnames: list[str] or str
        """
        kwargs["fieldnames"] = fieldnames.split() if isinstance(fieldnames, str) else fieldnames
        return self._xsv_import(
            xsv_source,
            encoding=encoding,
            transforms=transforms,
            filters=filters,
            row_class=row_class,
            limit=limit,
            delimiter="\t",
            **kwargs,
        )

    def _excel_import(
        self,
        excel_source: _ImportExportDataContainer,
        transforms: Optional[dict] = None,
        filters: Optional[dict] = None,
        row_class: Optional[type] = None,
        limit: Optional[int] = None,
        **kwargs: Any,
    ) -> Table:
        if openpyxl is None:
            raise Exception("openpyxl module not installed")

        def excel_as_dict(filename, **reader_args) -> Iterable[dict[str, str]]:
            with contextlib.closing(openpyxl.load_workbook(filename, read_only=True)) as wb:
                # read requested sheet if provided on kwargs, otherwise read active sheet
                requested_sheet = reader_args.get("sheet")
                ws = wb[requested_sheet] if requested_sheet else wb.active

                rows_iter = iter(ws.rows)

                # check whether to include or omit the header
                header = (reader_args.get("fieldnames")
                          or [str(cell.value) for cell in next(rows_iter)])

                for row in rows_iter:
                    yield {key: cell.value for key, cell in zip(header, row)}

        url_arg_names = "headers data username password cafile capath context".split()
        url_args = {k: kwargs.pop(k) for k in url_arg_names if k in kwargs}

        return self._import(
            excel_source,
            transforms=transforms,
            filters=filters,
            reader=lambda src: excel_as_dict(src._iterobj, **kwargs),
            row_class=row_class,
            limit=limit,
            url_args=url_args,
        )

    def excel_import(
        self,
        excel_source: _ImportExportDataContainer,
        transforms: Optional[dict] = None,
        filters: Optional[dict] = None,
        row_class: Optional[type] = None,
        limit: Optional[int] = None,
        fieldnames: Optional[Union[Iterable[str], str]] = None,
        **kwargs: Any,
    ) -> Table:
        """
        Imports the contents of a Excel file into this table.
        @param excel_source: Excel file - if a string is given, the file with that name will be
            opened, read, and closed; if a file object is given, then that object
            will be read as-is, and left for the caller to be closed.
        @type excel_source: string or file
        @param transforms: dict of functions by attribute name; if given, each
            attribute will be transformed using the corresponding transform; if there is no
            matching transform, the attribute will be read as a string (default); the
            transform function can also be defined as a (function, default-value) tuple; if
            there is an Exception raised by the transform function, then the attribute will
            be set to the given default value
        @type transforms: dict (optional)
        @param filters: dict of functions by attribute name; if given, each
            newly-read record will be filtered before being added to the table, with each
            filter function run using the corresponding attribute; if any filter function
            returns False, the record is not added to the table. Useful when reading large
            input files, to pre-screen only for data matching one or more filters
        @type filters: dict (optional)
        @param row_class: class to construct for each imported row when populating table (default=SimpleNamespace)
        @type row_class: type
        @param limit: number of records to import
        @type limit: int (optional)
        @param kwargs: additional arguments for the Excel reader. Only available argument is "sheet" to select which
            sheet to read (defaults to active sheet). kwargs may also contain any of the following if importing
            using a URL: C{headers}, C{data}, C{username}, C{password}
        @type kwargs: named arguments (optional)
        @param fieldnames: names for imported columns; used if there is no header line in the input file
        @type fieldnames: list[str] or str
        """
        kwargs["fieldnames"] = fieldnames.split() if isinstance(fieldnames, str) else fieldnames
        url_arg_names = "headers data username password".split()
        url_args = {k: kwargs.pop(k) for k in url_arg_names if k in kwargs}
        return self._excel_import(
            excel_source,
            transforms=transforms,
            filters=filters,
            row_class=row_class,
            limit=limit,
            url_args=url_args,
            **kwargs,
        )

    def csv_export(
        self,
        csv_dest: Optional[_ImportExportDataContainer] = None,
        fieldnames: Optional[Iterable[str]] = None,
        encoding: str = "utf-8",
        delimiter: str = ",",
        **kwargs: Any,
    ) -> Optional[str]:
        """
        Exports the contents of the table to a CSV-formatted file.
        @param csv_dest: CSV file - if a string is given, the file with that name will be
            opened, written, and closed; if a file object is given, then that object
            will be written as-is, and left for the caller to be closed.
            If None, then a string containing the exported data is returned.
        @type csv_dest: string or file
        @param fieldnames: attribute names to be exported; can be given as a single
            string with space-delimited names, or as a list of attribute names
        @type fieldnames: list of strings
        @param encoding: string (default="UTF-8"); if csv_dest is provided as a string
            representing an output filename, an encoding argument can be provided
        @type encoding: string
        @param delimiter: string (default=",") - overridable delimiter for value separator
        @type delimiter: string
        @param kwargs: additional keyword args to pass through to csv.DictWriter
        @type kwargs: named arguments (optional)

        If no destination file is given, the CSV-formatted data is returned as a string.
        """
        non_writer_args = "encoding csv_dest fieldnames".split()
        writer_args = {
            k: v for k, v in kwargs.items() if k not in non_writer_args
        }
        close_on_exit = False
        return_dest_value = False

        if csv_dest is None:
            csv_dest = io.StringIO()
            return_dest_value = True
        if isinstance(csv_dest, Path):
            csv_dest = str(csv_dest)
        if isinstance(csv_dest, str):
            csv_dest = open(csv_dest, "w", newline="", encoding=encoding)
            close_on_exit = True
        try:
            if fieldnames is None:
                fieldnames = self._attr_names()
            if isinstance(fieldnames, str):
                fieldnames = fieldnames.split()

            csv_dest.write(delimiter.join(fieldnames) + NL)
            csvout: csv.DictWriter = csv.DictWriter(
                csv_dest,
                list(fieldnames),
                extrasaction="ignore",
                lineterminator=NL,
                delimiter=delimiter,
                **writer_args,
            )
            try:
                csvout.writerows(_to_dict(o) for o in self.obs)
            except UnableToExtractAttributeNamesError:
                attr_fetch = attrgetter(*fieldnames)
                for o in self.obs:
                    csvout.writerow(dict(zip(fieldnames, attr_fetch(o))))
        finally:
            if close_on_exit:
                csv_dest.close()

        if return_dest_value:
            return csv_dest.getvalue()
        else:
            return None

    def tsv_export(
        self,
        tsv_dest: Optional[_ImportExportDataContainer],
        fieldnames: Optional[Iterable[str]] = None,
        encoding: str = "UTF-8",
        **kwargs: Any,
    ) -> Optional[str]:
        r"""
        Similar to csv_export, with delimiter="\t"
        """
        return self.csv_export(
            tsv_dest, fieldnames=fieldnames, encoding=encoding, delimiter="\t", **kwargs
        )

    def json_import(
        self,
        source: _ImportExportDataContainer,
        encoding: str = "UTF-8",
        transforms: Optional[dict] = None,
        row_class: Optional[type] = None,
        streaming: bool = False,
        path: str = "",
        json_decoder: Optional[json.JSONDecoder] = None,
        **kwargs: Any,
    ) -> Table:
        """
        Imports the contents of a JSON data file into this table.
        @param source: JSON data file - if a string is given, the file with that name will be
            opened, read, and closed; if a file object is given, then that object
            will be read as-is, and left for the caller to be closed.
        @type source: string or file
        @param transforms: dict of functions by attribute name; if given, each
            attribute will be transformed using the corresponding transform; if there is no
            matching transform, the attribute will be read as a string (default); the
            transform function can also be defined as a (function, default-value) tuple; if
            there is an Exception raised by the transform function, then the attribute will
            be set to the given default value
        @type transforms: dict (optional)
        @param row_class: class to construct for each imported row when populating table (default=SimpleNamespace)
        @type row_class: type
        @param streaming: boolean flag to indicate whether inbound JSON will be a stream of multiple objects
            or a single list object (default=False)
        @type streaming: bool
        @param path: (only valid if streaming=False) a '.'-delimited path into the inbound JSON, in case
            the objects to import are not in a top-level JSON list
        @type path: str
        @param json_decoder: subclass of json.JSONDecoder to pass through to json.loads (default=None)
        @type json_decoder: json.JSONDecoder
        """
        class PathNotFoundError(KeyError):
            def __init__(self, key, nf_path):
                super().__init__(key)
                self.key = key
                self.path = nf_path

            def __str__(self):
                return f"could not find {self.key!r} element of path {self.path!r} in imported JSON"

        class _JsonFileReader:
            def __init__(self, src):
                self.source = src
                self.streaming = streaming

            def __iter__(self):
                if self.streaming:
                    # incrementally read lines from source until a valid JSON object
                    # can be parsed
                    current = ""
                    for line in self.source:
                        if current:
                            current += " "
                        current += line
                        try:
                            yield json.loads(current, cls=json_decoder)
                            current = ""
                        except Exception:  # noqa
                            pass
                else:
                    # merge entire source into one JSON parseable object
                    inbound_json = '\n'.join(self.source)
                    obs = json.loads(inbound_json, cls=json_decoder)

                    # descend into parsed JSON object by path
                    for path_item in filter(None, path.split(".")):
                        obs = obs.get(path_item)
                        if obs is None:
                            raise PathNotFoundError(path_item, path)

                    yield from obs

        if path and streaming:
            raise ValueError("cannot specify path and streaming=True")

        if row_class is None:
            row_class = default_row_class

        url_arg_names = "headers data username password cafile capath context".split()
        url_args = {k: kwargs.pop(k) for k in url_arg_names if k in kwargs}

        return self._import(
            source,
            encoding,
            transforms=transforms,
            reader=_JsonFileReader,
            row_class=row_class,
            url_args=url_args,
        )

    def json_export(
        self,
        dest: Optional[_ImportExportDataContainer] = None,
        fieldnames: Optional[Union[Iterable[str], str]] = None,
        encoding: str = "UTF-8",
        streaming: bool = False,
        json_encoder: Optional[Union[type[json.JSONEncoder], tuple[type[json.JSONEncoder], ...]]] = None,
    ) -> Optional[str]:
        """
        Exports the contents of the table to a JSON-formatted file.
        @param dest: output file - if a string is given, the file with that name will be
            opened, written, and closed; if a file object is given, then that object
            will be written as-is, and left for the caller to be closed.
            If None, then a string containing the exported data is returned.
        @type dest: string or file
        @param fieldnames: attribute names to be exported; can be given as a single
            string with space-delimited names, or as a list of attribute names
        @type fieldnames: list of strings
        @param encoding: string (default="UTF-8"); if csv_dest is provided as a string
            representing an output filename, an encoding argument can be provided
        @type encoding: string
        @param streaming: bool (default=False); flag to return JSON as a separate
            JSON string for each object in the table. If False, returns a single
            JSON list containing all the table objects.
        @type streaming: bool
        @param json_encoder: an encoder or tuple of encoders to perform custom
            JSON encoding for fields in table objects
        @type json_encoder: json.JSONEncoder or tuple(json.JSONEncoder)
        """
        close_on_exit = False
        return_dest_value = False

        if json_encoder is not None:
            if isinstance(json_encoder, tuple):
                # use multiple inheritance to chain encoders
                json_encoder = type(
                    "compound_JSON_encoder",
                    (*json_encoder, json.JSONEncoder),
                    {}
                )
            json_encoder = cast(json.JSONEncoder, json_encoder)

        if dest is None:
            dest = io.StringIO()
            return_dest_value = True
        if isinstance(dest, Path):
            dest = str(Path)
        if isinstance(dest, str):
            dest = open(dest, "w", encoding=encoding)
            close_on_exit = True
        try:
            if isinstance(fieldnames, str):
                fieldnames = fieldnames.split()

            if streaming:
                if fieldnames is None:
                    for o in self.obs:
                        dest.write(_to_json(o, json_encoder) + "\n")
                else:
                    for o in self.obs:
                        dest.write(
                            json.dumps({f: getattr(o, f, None) for f in fieldnames},
                                       cls=json_encoder) + "\n"
                        )
            else:
                dest.write("[\n")
                if self.obs:
                    if fieldnames is None:
                        for o in itertools.islice(self.obs, len(self.obs)-1):
                            dest.write(_to_json(o, json_encoder) + ",\n")
                        o = self.obs[-1]
                        dest.write(_to_json(o, json_encoder) + "\n")
                    else:
                        for o in itertools.islice(self.obs, len(self.obs)-1):
                            dest.write(
                                json.dumps({f: getattr(o, f, None) for f in fieldnames},
                                           cls=json_encoder) + ",\n"
                            )
                        o = self.obs[-1]
                        dest.write(
                            json.dumps({f: getattr(o, f, None) for f in fieldnames},
                                       cls=json_encoder) + "\n"
                        )
                dest.write("]\n")

        finally:
            if close_on_exit:
                dest.close()

        if return_dest_value:
            return dest.getvalue()
        else:
            return None

    def excel_export(
        self,
        excel_dest: _ImportExportDataContainer,
        fieldnames: Optional[Iterable[str]] = None,
        **kwargs: Any,
    ):
        """
        Exports the contents of the table to an Excel .xslx file.
        @param excel_dest: Excel file - if a string is given, the file with that name will be
            opened, written, and closed; if a file object is given, then that object
            will be written as-is, and left for the caller to be closed.
        @type excel_dest: string or file
        @param fieldnames: attribute names to be exported; can be given as a single
            string with space-delimited names, or as a list of attribute names
        @type fieldnames: list of strings
        @param kwargs: additional keyword args
        @type kwargs: named arguments (optional)
        """
        if openpyxl is None:
            raise Exception("openpyxl module not installed")
        if kwargs.pop('lxml', True) is False:
            lxml = None
        else:
            try:
                import lxml
            except ImportError:
                lxml = None

        # lxml enables write_only mode (which is faster)
        if lxml is not None:
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet()
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
        # set header rows
        if fieldnames is None:
            fieldnames = self._attr_names()
        elif isinstance(fieldnames, str):
            fieldnames = fieldnames.split()
        ws.append(fieldnames)
        # append data
        for o in self.obs:
            ws.append([v for v in _to_dict(o).values()])
        wb.save(excel_dest)

    def as_dataframe(self, fields: Optional[Union[Iterable[str], str]] = None):
        """
        Export contents of the Table to a pandas DataFrame.
        @param fields: list of strings, or single space-delimited string, listing
        attribute name to be included in the output
         - names starting with '-' indicate to suppress that field
         - '*' means include all other field names
         - if no fields are specifically included, then all fields are used
        @type fields: list, or space-delimited string
        """
        try:
            import pandas as pd
        except ImportError:
            print("pandas not installed", file=sys.stderr)
            return None

        if fields is None:
            # assume all fields
            fieldnames = self._attr_names()
        else:
            fieldnames = self._parse_fields_string(fields)

        # attrgetter is an efficient extractor of values from objects
        extractor = attrgetter(*fieldnames)

        # build a DataFrame from the tuples returned by the extractor for
        # each object in this Table
        ret = pd.DataFrame(
            map(extractor, self),
            columns=fieldnames,
        )
        return ret

    def compute_field(
        self, attrname: str, fn: Union[Callable[[Any], Any], str], default: Optional[Any] = None
    ) -> Table:
        """
        Computes a new attribute for each object in table, or replaces an
        existing attribute in each record with a computed value
        @param attrname: attribute to compute for each object
        @type attrname: string
        @param fn: function used to compute new attribute value, based on
        other values in the object, as in::

            lambda ob : ob.commission_pct/100.0 * ob.gross_sales

        @type fn: function(obj) returns value
        @param default: value to use if an exception is raised while trying
        to evaluate fn

        fn can also be passed as just a string, referencing an existing field
        that might be difficult to reference (if it has embedded spaces or
        punctuation).
        """
        if attrname in self._indexes:
            idx = self._indexes[attrname]
            idx._clear()
            idx_setitem = idx.__setitem__
        else:
            # there is no index for this attr
            idx_setitem = None

        if isinstance(fn, str):
            fn = lambda r, attrname=fn: getattr(r, attrname, None)

        try:
            for rec_ in self:
                try:
                    val = fn(rec_)
                except Exception:  # noqa
                    val = default
                if isinstance(rec_, DataObject):
                    rec_.__dict__[attrname] = val
                else:
                    setattr(rec_, attrname, val)
                # update index for this attribute, if there is one
                if idx_setitem is not None:
                    idx_setitem(val, rec_)
        except AttributeError:
            raise AttributeError(
                f"cannot add/modify attribute {attrname!r} in table records"
            )

        self._contents_changed(invalidate_search_indexes=False)
        return self

    add_field = compute_field

    def groupby_with_summaries(self, keyexpr, **outexprs):
        """
        simple prototype of group by, with support for expressions in the group-by clause
        and outputs
        @param keyexpr: grouping field and optional expression for computing the key value
        @type keyexpr: string or tuple
        @param outexprs: named arguments describing one or more summary values to
        compute per key
        @type outexprs: callable, taking a sequence of objects as input and returning
        a single summary value
        """
        if isinstance(keyexpr, str):
            keyattrs = keyexpr.split()
            keyfn = attrgetter(*keyattrs)

        elif isinstance(keyexpr, tuple):
            keyattrs = (keyexpr[0],)
            keyfn = keyexpr[1]

        else:
            raise TypeError("keyexpr must be string or tuple")

        grouped_obs = defaultdict(list)
        for ob in self.obs:
            grouped_obs[keyfn(ob)].append(ob)

        tbl = Table()
        for k in keyattrs:
            tbl.create_index(k, unique=(len(keyattrs) == 1))
        for key, recs in sorted(grouped_obs.items()):
            group_obj = default_row_class(**dict(zip(keyattrs, key)))
            for subkey, expr in outexprs.items():
                setattr(group_obj, subkey, expr(recs))
            tbl.insert(group_obj)
        return tbl

    def groupby(
            self,
            keyexpr: Union[str, List[str], Callable[[TableContent], Any]],
            sort: bool = False
    ) -> Iterable[Tuple[Any, Table[TableContent]]]:
        """
        Analogous to itertools.groupby, using the Table as the iterable to be
        grouped.
        @param keyexpr: grouping field(s) or function for computing the key value
        @type keyexpr: string, list of strings, or callable
        @param sort: flag indicating whether the Table should be sorted before
        grouping
        @type sort: bool, default=False
        """
        if isinstance(keyexpr, list):
            keyfn = attrgetter(keyexpr[0], *keyexpr[1:])

        elif isinstance(keyexpr, Callable):
            keyfn = keyexpr

        elif isinstance(keyexpr, str):
            keyfn = lambda o: getattr(o, keyexpr, None)

        else:
            raise TypeError("keyexpr must be string, list of strings, or callable")

        if sort:
            self.sort(keyfn)

        for key, group_objs in itertools.groupby(self.obs, keyfn):
            group_tbl = self.copy_template().insert_many(group_objs)
            yield key, group_tbl

    def batched(self, batch_size: int) -> Iterable[Table[TableContent]]:
        """
        Yields subtables of size `batch_size`, such as might be returned in
        a paginated API. The final Table may be less than `batch_size` in size.
        """
        if batch_size <= 0:
            raise ValueError("batch_size must be greater than 0")

        num_items = len(self)
        offset = 0
        while offset < num_items:
            yield self[offset: offset + batch_size]
            offset += batch_size

    def splitby(
            self,
            pred: Union[str, PredicateFunction] = None,
            *,
            errors: Union[bool, str, dict[type[Exception], Union[bool, str]]] = "discard",
            **kwargs,
    ) -> tuple[Table[TableContent], ...]:
        """
        Takes a predicate function (takes a table record and returns True or False)
        and returns two tables: a table with all the rows that returned False and
        a table with all the rows that returned True. Will also accept a string
        indicating a particular field name, and uses `bool(getattr(rec, field_name))`
        for the predicate function.

              is_odd = lambda x: bool(x % 2)
              evens, odds = tbl.splitby(lambda rec: is_odd(rec.value))
              nulls, not_nulls = tbl.splitby("optional_data_field")

        A shorthand for specifying a predicate function can be used if the predicate
        is solely a test for a specific value for one or more attributes:

              qa_data, production_assembly_data = data.splitby(
                lambda rec: rec.env == "prod" and rec.dept == "assembly"
              )
              # can be written as
              qa_data, production_data = data.splitby(env="prod", dept="assembly")

        An optional `errors` argument can be passed to define what action to take
        if an exception occurs while evaluating the predicate function. Valid values
        for `errors` are:
            True : return exceptions as True
            False: return exceptions as False
            'discard': do not return table rows that raise exceptions
            'return': return a third table containing rows that raise exceptions
            'raise': raise the exception

        `errors` can also be given as a dict mapping Exception types to one of
        these 4 values.

        The default value for `errors` (if omitted, or if an exception is raised
        that is not listed in `errors`), is to discard the row.
        """

        # validate pred and kwargs args; if kwargs specified, synthesize pred
        # function from them
        if pred is None:
            if not kwargs:
                raise ValueError(
                    "must provide either a predicate function or one or more named"
                    " arguments using table field names and splitting values"
                )
            if len(kwargs) > 1:
                getvalues = attrgetter(*kwargs.keys())
                matchvalues = tuple(kwargs.values())
                pred = (
                    lambda split_rec, gv=getvalues, mv=matchvalues: gv(split_rec) == mv
                )
            else:
                key, value = next(iter(kwargs.items()))
                pred = lambda split_rec: getattr(split_rec, key, None) == value
        else:
            if kwargs:
                raise ValueError(
                    "must provide either a predicate function or one or more named"
                    " arguments, not both"
                )

        # if key is a str, convert it to a predicate function using getattr
        if isinstance(pred, str):
            key_str = pred
            pred = lambda obj: getattr(obj, key_str, None)

        discard_errors = object()
        reraise_errors = object()
        RETURN_ERRORS_TABLE = 2
        if errors is None:
            error_responses = {}
        elif isinstance(errors, bool):
            error_responses = {Exception: errors}
        elif errors == "return":
            error_responses = {Exception: RETURN_ERRORS_TABLE}
        elif errors == "discard":
            error_responses = {Exception: discard_errors}
        elif errors == "raise":
            error_responses = {Exception: reraise_errors}
        elif isinstance(errors, dict):
            if not all(issubclass(error_exception, Exception) for error_exception in errors):
                raise ValueError(
                    f"one or more error exception types is invalid {errors!r};"
                    " must be Exception or a subclass of Exception"
                )

            if not all(
                    error_response in {True, False, "return", "discard", "raise"}
                    for error_response in errors.values()
            ):
                raise ValueError(
                    f"one or more error values is invalid {errors!r};"
                    " must be True, False, 'return', or 'discard'"
                )
            error_responses = {
                k: {
                    "discard": discard_errors,
                    "return": RETURN_ERRORS_TABLE,
                    "raise": reraise_errors,
                }.get(v, v) for k, v in errors.items()
            }
        else:
            raise ValueError(
                f"invalid type/value for errors: {errors!r};"
                " must be True, False, 'return', or 'discards',"
                " or a dict mapping Exception types to one of those values")

        # wrap pred in try-except, to infer any failure of pred -> False,
        # and use not not to bool-ify the value returned from pred()
        def wrapped_pred(obj):
            try:
                return not not pred(obj)
            except Exception as exc:
                for exctype, retval in error_responses.items():
                    if isinstance(exc, exctype):
                        return retval
                    if retval is reraise_errors:
                        raise
                return discard_errors

        # construct return tables to receive False and True evaluated records
        ret = self.copy_template(), self.copy_template()
        if any(
                error_response is RETURN_ERRORS_TABLE
                for error_response in error_responses.values()
        ):
            ret += (self.copy_template(),)

        # iterate over self and evaluate predicate for each record - use groupby to take
        # advantage of efficiencies when using insert_many() over multiple insert() calls
        for pred_value, recs in itertools.groupby(self, key=wrapped_pred):
            if pred_value is discard_errors:
                continue
            ret[pred_value].insert_many(recs)

        return ret

    def unique(self, key: Optional[Union[Callable[[TableContent], Any], str]] = None) -> Table[TableContent]:
        """
        Create a new table of objects,containing no duplicate values.

        @param key: (default=None) optional callable for computing a representative unique key for each
        object in the table. If None, then a key will be composed as a tuple of all the values in the object.
        @type key: callable, takes the record as an argument, and returns the key value or tuple to be used
        to represent uniqueness.
        """
        if isinstance(key, str):
            key = cast(Callable[[TableContent], Any], lambda r, attr=key: getattr(r, attr, None))
        ret = self.copy_template()
        seen = set()
        for ob in self:
            if key is None:
                ob_dict = _to_dict(ob)
                reckey = tuple(sorted(ob_dict.items()))
            else:
                reckey = key(ob)
            if reckey not in seen:
                seen.add(reckey)
                ret.insert(ob)
        return ret

    def info(self) -> dict[str, Any]:
        """
        Quick method to list informative table statistics
        :return: dict listing table information and statistics
        """
        unique_indexes = set(self._uniqueIndexes)
        return {
            "len": len(self),
            "name": self.table_name,
            "fields": self._attr_names(),
            "indexes": [
                (idx_name, self._indexes[idx_name] in unique_indexes)
                for idx_name in self._indexes
            ],
            "created": self.create_time,
            "modified": self.modify_time,
            "last_import": self.import_time,
        }

    def head(self, n: int = 10) -> Table[TableContent]:
        """
        Return a table of the first 'n' records in a table.
        :param n: (int, default=10) number of records to return
        :return: Table
        """
        return self[:n](self.table_name)

    def tail(self, n: int = 10) -> Table[TableContent]:
        """
        Return a table of the last 'n' records in a table.
        :param n: (int, default=10) number of records to return
        :return: Table
        """
        return self[-n:](self.table_name)

    def stats(self, field_names: Optional[Union[str, Iterable[str]]] = None, by_field: bool = True) -> Table:
        """
        Return a summary Table of statistics for numeric data in a Table.
        For each field in the source table, returns:
        - mean
        - min
        - max
        - variance
        - standard deviation
        - count
        - missing
        :param field_names:
        :param by_field:
        :return: Table of statistics; if by_field=True, each row contains summary
                 statistics for each field; if by_field =False, each row contains a
                 statistic and the value of that statistic for each field (conceptually
                 a transpose of the by_field=True results)
        """
        ret: Table = Table()

        # if table is empty, return empty stats
        if not self:
            return ret

        if field_names is None:
            field_names = self._parse_fields_string("*")

        accum = {
            fname: list(
                filter(lambda x: isinstance(x, _numeric_type), getattr(self.all, fname))
            )
            for fname in field_names
        }

        def safe_fn(fn, seq):
            try:
                return fn(seq)
            except Exception:  # noqa
                return None

        def rounding(fn, x):
            import math
            v = fn(x)
            if v in (None, 0.0):
                return v
            if abs(v) > 1:
                mag = int(math.log10(abs(v))) + 1
                return round(v, max(4-mag, 0))
            return v

        stat_fn_map = (
            ("mean", partial(rounding, partial(safe_fn, getattr(statistics, "fmean", statistics.mean)))),
            ("min", partial(safe_fn, min)),
            ("max", partial(safe_fn, max)),
            ("variance", partial(rounding, partial(safe_fn, statistics.variance))),
            ("std_dev", partial(rounding, partial(safe_fn, statistics.stdev))),
            ("count", len),
            ("missing", lambda seq: len(self) - len(seq)),
        )

        if by_field:
            ret.create_index("name", unique=True)
            ret.insert_many(default_row_class(name=fname,
                                              **{stat_name: stat_fn(accum[fname])
                                                 for stat_name, stat_fn in stat_fn_map})
                            for fname in field_names)
        else:
            ret.create_index("stat", unique=True)
            ret.insert_many(default_row_class(stat=stat_name,
                                              **{fname: stat_fn(accum[fname])
                                                 for fname in field_names})
                            for stat_name, stat_fn in stat_fn_map)
        return ret

    def _parse_fields_string(self, field_names: Union[str, Iterable[str]]) -> list[str]:
        """
        Convert raw string or list of names to actual column names:
        - names starting with '-' indicate to suppress that field
        - '*' means include all other field names
        - if no fields are specifically included, then all fields are used
        :param field_names: str or list
        :return: expanded list of field names
        """
        if isinstance(field_names, str):
            field_names = field_names.split()
        if not self.obs:
            return list(field_names)

        suppress_names = [nm[1:] for nm in field_names if nm.startswith("-")]
        field_names = [nm for nm in field_names if not nm.startswith("-")]
        if not field_names:
            field_names = ["*"]
        if "*" in field_names:
            if self:
                star_fields = [
                    name for name in self._attr_names() if name not in field_names
                ]
            else:
                # no records to look at, just use names of any defined indexes
                star_fields = list(self._indexes.keys())
            fn_iter = iter(field_names)
            field_names = (
                list(itertools.takewhile(lambda x: x != "*", fn_iter))
                + star_fields
                + list(fn_iter)
            )
        field_names = [nm for nm in field_names if nm not in suppress_names]
        return field_names

    def _rich_table(
            self,
            fields: Optional[Iterable[Union[str, dict]]] = None,
            empty: Any = "",
            groupby: Optional[str] = None,
            **kwargs
    ):
        if rich is None:
            raise Exception("rich module not installed")

        from rich.table import Table as RichTable

        if fields is None:
            fields = self.info()["fields"]

        attr_names: list[str] = []
        field_settings: list[tuple[str, dict]] = []

        for field_spec in fields:
            if isinstance(field_spec, str):
                name, field_spec = field_spec, {}
                # find a value for this attribute, and if numeric, make column right-justified
                next_v = next(
                    (v for v in getattr(self.all, name) if v is not None), None
                )
                if isinstance(next_v, right_justify_types):
                    field_spec["justify"] = "right"
                else:
                    try:
                        if all(
                            len(str(v)) <= 1
                            for v in getattr(self.all, name).unique
                            if v is not None
                        ):
                            field_spec["justify"] = "center"
                    except TypeError:
                        pass
            else:
                # use field settings form caller
                name, field_spec = field_spec

            attr_names.append(name)
            header = field_spec.pop("header", None)
            if header is None:
                header = name.replace("_", " ").title() if name.islower() else name
            field_settings.append((header, field_spec))

        grouping = False
        group_attrs: list[str] = []
        if groupby is not None:
            group_attrs = [g for g in self._parse_fields_string(groupby)
                           if g in attr_names]
            if group_attrs:
                grouping = True

        table_defaults = dict(show_header=True, header_style="bold", box=box.ASCII)
        if getattr(sys.stdout, "isatty", lambda: False)():
            table_defaults["box"] = box.SIMPLE
        if self.table_name:
            table_defaults["title"] = self.table_name
        table_kwargs = table_defaults
        table_kwargs.update(kwargs)

        # create rich Table
        rt = RichTable(**table_kwargs)

        # define rich Table columns
        for header, field_spec in field_settings:
            rt.add_column(header, **field_spec)

        # add row data from self to rich Table
        if not grouping:
            for row in self.formatted_table(*fields):
                rt.add_row(*[getattr(row, attr_name, empty) for attr_name in attr_names])
        else:
            prev = ("",) * len(group_attrs)
            determine_suppressed_attrs = _determine_suppressed_attrs
            for row in self.formatted_table(*fields):
                curr = tuple(getattr(row, attr, "") for attr in group_attrs)
                suppress_attrs = determine_suppressed_attrs(group_attrs, prev, curr)
                row_items = [
                    "" if attr_name in suppress_attrs else getattr(row, attr_name, empty)
                    for attr_name in attr_names
                ]
                rt.add_row(*row_items)
                prev = curr

        return rt

    def present(
            self,
            fields: Optional[Iterable[str]] = None,
            file: Optional[TextIO] = None,
            groupby: Optional[str] = None,
            **kwargs: Any
    ) -> None:
        """
        Print a nicely-formatted table of the records in the Table, using the `rich`
        Python module. If the Table has a title, then that will be displayed as the
        title over the tabular output.

        :param fields: list of field names to include in the tabular output
        :param file: (optional) output file for tabular output (defaults to sys.stdout)
        :param groupby: (optional) field name or space-delimited list of field names
                        for groups to be indicated by suppressing consecutive duplicate
                        values in a column
        :param kwargs: (optional) additional keyword args to customize the `rich` output,
                       as might be passed to the `rich.Table` class
        :return: None

        Note: the `rich` Python module must be installed to use this method.
        """
        try:
            from rich.console import Console
        except ImportError:
            raise Exception("rich module not installed")

        console = Console(file=file)
        table_kwargs = {"header_style": "bold yellow"}
        table_kwargs.update(kwargs)
        table = self._rich_table(fields, empty="", groupby=groupby, **table_kwargs)
        print()
        console.print(table)

    def as_html(
            self,
            fields: Union[str, Iterable[str]] = "*",
            formats: Optional[dict[str, str]] = None,
            groupby: Optional[Union[str, Iterable[str]]] = None,
            table_properties: Optional[dict] = None,
    ) -> str:
        """
        Output the table as a rudimentary HTML table.
        @param fields: fields in the table to be shown in the table
                       - listing '*' as a field will add all unnamed fields
                       - starting a field name with '-' will suppress that name
        @type fields: list of strings or a single space-delimited string
        @param formats: optional dict of str formats to use when converting field values
                        to strings (usually used for float conversions, but could also be
                        used for str conversion or text wrapping)
        @type formats: mapping of field names or types to either str formats as used by
                       the str.format method, or a callable that takes a value and returns
                       a str
        @param groupby: optional field name for groups to be indicated by suppressing
                        consecutive duplicate values in a column
        @type groupby: str
        @return: string of generated HTML representing the selected table row attributes
        """
        fields = self._parse_fields_string(fields)
        if formats is None:
            formats = {}
        field_format_map = {}
        attr_names = fields

        def row_to_tr(r: TableContent, suppress: Iterable[str] = ()) -> str:
            ret_tr = ["<tr>"]
            for fld in fields:
                align = "left"
                if fld not in suppress:
                    v = getattr(r, fld, "")
                    if isinstance(v, right_justify_types):
                        align = "right"
                    if fld not in field_format_map:
                        field_format_map[fld] = formats.get(fld, formats.get(type(v), "{}"))
                    v_format = field_format_map[fld]
                    str_v = v_format.format(v) if isinstance(v_format, str) else v_format(v)
                else:
                    str_v = ""
                ret_tr.append(f'<td><div align="{align}">{str_v}</div></td>')
            ret_tr.append("</tr>\n")
            return "".join(ret_tr)

        grouping = False
        group_attrs: list[str] = []
        if groupby is not None:
            group_attrs = [g for g in self._parse_fields_string(groupby)
                           if g in attr_names]
            if group_attrs:
                grouping = True

        if table_properties is not None:
            table_modifiers = "".join(f' {k}="{v}"' for k, v in table_properties.items())
        else:
            table_modifiers = ""

        headers = "".join(f'<th><div align="center">{fld}</div></th>' for fld in fields)
        if not grouping:
            ret = (
                f"<table{table_modifiers}>\n<thead>\n"
                f"<tr>{headers}</tr>\n"
                "</thead>\n<tbody>"
                f"{''.join(row_to_tr(row) for row in self)}"
                "</tbody>\n</table>"
            )
        else:
            rows: list[str] = []
            prev = ("",) * len(group_attrs)
            determine_suppressed_attrs = _determine_suppressed_attrs
            for row in self:
                curr = tuple(getattr(row, attr, "") for attr in group_attrs)
                suppress_attrs = determine_suppressed_attrs(group_attrs, prev, curr)
                rows.append(row_to_tr(row, suppress=suppress_attrs))
                prev = curr
            ret = (
                f"<table{table_modifiers}>\n<thead>\n"
                f"<tr>{headers}</tr>\n"
                "</thead>\n<tbody>"
                f"{''.join(rows)}"
                "</tbody>\n</table>"
            )
        return ret

    def as_markdown(
            self,
            fields: Union[str, Iterable[str]] = "*",
            formats: Optional[dict[str, str]] = None,
            groupby: Optional[Union[str, Iterable[str]]] = None,
    ) -> str:
        """
        Output the table as a Markdown table.
        @param fields: fields in the table to be shown in the table
                       - listing '*' as a field will add all unnamed fields
                       - starting a field name with '-' will suppress that name
        @type fields: list of strings or a single space-delimited string
        @param formats: optional dict of str formats to use when converting field values
                        to strings (usually used for float conversions, but could also be
                        used for str conversion or text wrapping
        @type formats: mapping of field names or types to either str formats as used by
                       the str.format method, or a callable that takes a value and returns
                       a str
        @param groupby: optional field name for groups to be indicated by suppressing
                        consecutive duplicate values in a column
        @type groupby: str
        @return: string of generated Markdown representing the selected table row attributes
        """
        fields = self._parse_fields_string(fields)
        if formats is None:
            formats = {}
        field_format_map = {}
        attr_names = fields

        grouping = False
        group_attrs: list[str] = []
        if groupby is not None:
            group_attrs = [g for g in self._parse_fields_string(groupby)
                           if g in attr_names]
            if group_attrs:
                grouping = True

        center_vals = (True, False, 'Y', 'N', 'X', 'YES', 'NO', 'y', 'n', 'x', 'yes', 'no', 0, 1, None)
        field_align_map = {}
        for f in fields:
            align = "---"
            align_center = True
            align_right = True
            for v in getattr(self.all, f):
                if align_center and v in center_vals:
                    continue
                align_center = False
                if not (v is None or isinstance(v, right_justify_types)):
                    align_right = False
                if not align_right and not align_center:
                    break
            if align_center:
                align = ":---:"
            elif align_right:
                align = "---:"
            field_align_map[f] = align

        def row_to_tr(r: TableContent, suppress: Iterable[str] = ()) -> str:
            ret_tr = ["|"]
            for fld in fields:
                if fld not in suppress:
                    fld_v = getattr(r, fld, "")
                    if fld not in field_format_map:
                        field_format_map[fld] = formats.get(fld, formats.get(type(fld_v), "{}"))
                    v_format = field_format_map[fld]
                    str_v = v_format.format(fld_v) if isinstance(v_format, str) else v_format(fld_v)
                else:
                    str_v = ""
                ret_tr.append(f" {str_v} |")
            ret_tr.append("\n")
            return "".join(ret_tr)

        if not grouping:
            ret = (
                f"| {' | '.join(fields)} |\n"
                f"|{'|'.join(field_align_map[f] for f in fields)}|\n"
                f"{''.join(row_to_tr(row) for row in self)}"
            )
        else:
            rows: list[str] = []
            prev = ("",) * len(group_attrs)
            determine_suppressed_attrs = _determine_suppressed_attrs
            for row in self:
                curr = tuple(getattr(row, attr, "") for attr in group_attrs)
                suppress_attrs = determine_suppressed_attrs(group_attrs, prev, curr)
                rows.append(row_to_tr(row, suppress=suppress_attrs))
                prev = curr
            ret = (
                f"| {' | '.join(fields)} |\n"
                f"|{'|'.join(field_align_map[f] for f in fields)}|\n"
                f"{''.join(rows)}"
            )

        return ret


Sequence.register(Table)


# module-level convenience functions for Table.*_import() instance methods
def _make_module_level_import_fn(name: str) -> Callable:
    table_method = getattr(Table, name)

    @functools.wraps(table_method)
    def import_fn(*args, **kwargs):
        ret = Table()
        return table_method(ret, *args, **kwargs)

    return import_fn


csv_import = _make_module_level_import_fn("csv_import")
tsv_import = _make_module_level_import_fn("tsv_import")
json_import = _make_module_level_import_fn("json_import")
excel_import = _make_module_level_import_fn("excel_import")


class _PivotTable(Table):
    """Enhanced Table containing pivot results from calling table.pivot()."""

    def __init__(
            self,
            parent: Union[Table, _PivotTable],
            attr_val_path: list[tuple[str, str]],
            attrlist: Iterable[str],
    ):
        """PivotTable initializer - do not create these directly, use
        L{Table.pivot}.
        """
        super().__init__()
        self._attr_path: list[tuple[str, str]] = attr_val_path[:]
        self._pivot_attrs: list[str] = list(attrlist)
        self._subtable_dict: dict[str, _PivotTable] = {}
        self.subtables: list[_PivotTable] = []

        # for k,v in parent._indexes.items():
        #     self._indexes[k] = v.copy_template()
        self._indexes.update(
            {k: v.copy_template() for k, v in parent._indexes.items()}
        )
        if not attr_val_path:
            self.insert_many(parent.obs)
        else:
            attr, val = attr_val_path[-1]
            self.insert_many(parent.where(**{attr: val}))
            parent._subtable_dict[val] = self

        if len(self._pivot_attrs) > 0:
            this_attr, *sub_attrlist = self._pivot_attrs
            ind = parent._indexes[this_attr]
            self.subtables = [
                _PivotTable(self, attr_val_path + [(this_attr, k)], sub_attrlist)
                for k in sorted(ind.keys())
            ]

    def __getitem__(self, val):
        if self._subtable_dict:
            return self._subtable_dict[val]
        else:
            return super().__getitem__(val)

    def keys(self) -> Iterable[str]:
        return sorted(self._subtable_dict.keys())

    def items(self) -> Iterable[tuple[str, Any]]:
        return sorted(self._subtable_dict.items())

    def values(self) -> Iterable[Any]:
        return [self._subtable_dict[k] for k in self.keys()]

    def pivot_key(self) -> list[tuple[str, str]]:
        """
        Return the set of attribute-value pairs that define the contents of this
        table within the original source table.
        """
        return self._attr_path

    def pivot_key_str(self) -> str:
        """Return the pivot_key as a displayable string."""
        return "/".join(f"{attr}:{key}" for attr, key in self._attr_path)

    def has_subtables(self) -> bool:
        """Return whether this table has further subtables."""
        return bool(self.subtables)

    def dump(
            self,
            out: TextIO = sys.stdout,
            row_fn: Callable[[Any], str] = repr,
            limit: int = -1,
            indent: int = 0,
    ) -> None:
        """
        Dump out the contents of this table in a nested listing.
        @param out: output stream to write to
        @param row_fn: function to call to display individual rows
        @param limit: number of records to show at deepest level of pivot (-1=show all)
        @param indent: current nesting level
        """
        if indent:
            out.write("  " * indent + self.pivot_key_str())
        else:
            out.write(f"Pivot: {','.join(self._pivot_attrs)}")
        out.write(NL)
        if self.has_subtables():
            for sub in self.subtables:
                if sub:
                    sub.dump(out, row_fn, limit, indent + 1)
        else:
            if limit >= 0:
                showslice = slice(0, limit)
            else:
                showslice = slice(None, None)
            for r in self.obs[showslice]:
                out.write("  " * (indent + 1) + row_fn(r) + NL)
        out.flush()

    def dump_counts(
            self,
            out: TextIO = sys.stdout,
            count_fn: Callable[[Iterable[Any]], int] = len,
            colwidth: int = 10
    ) -> None:
        """
        Dump out the summary counts of entries in this pivot table as a tabular listing.
        @param out: output stream to write to
        @param count_fn: (default=len) function for computing value for each pivot cell
        @param colwidth: (default=10)
        """
        if len(self._pivot_attrs) == 1:
            out.write(f"Pivot: {','.join(self._pivot_attrs)}\n")
            maxkeylen = max(len(str(k)) for k in self.keys())
            maxvallen = colwidth
            keytally = {}
            for k, sub in self.items():
                sub_v = count_fn(sub)
                maxvallen = max(maxvallen, len(str(sub_v)))
                keytally[k] = sub_v
            for k, sub in self.items():
                out.write(
                    f"{str(k):<{maxkeylen}.{maxkeylen}s} {keytally[k]:>{maxvallen}}\n"
                )
        elif len(self._pivot_attrs) == 2:
            out.write(f"Pivot: {','.join(self._pivot_attrs)}\n")
            maxkeylen = max(max(len(str(k)) for k in self.keys()), 5)
            maxvallen = max(
                max(len(str(k)) for k in self.subtables[0].keys()), colwidth
            )
            keytally = dict.fromkeys(self.subtables[0].keys(), 0)
            out.write(f"{' ' * maxkeylen} ")
            out.write(
                " ".join(
                    f"{str(k):>{maxvallen}.{maxvallen}s}"
                    for k in self.subtables[0].keys()
                )
            )
            out.write(f' {"Total":>{maxvallen}s}\n')

            for k, sub in self.items():
                out.write(f"{str(k):<{maxkeylen}.{maxkeylen}s} ")
                for kk, ssub in sub.items():
                    ssub_v = count_fn(ssub)
                    out.write(f"{ssub_v:{maxvallen}d} ")
                    keytally[kk] += ssub_v
                    maxvallen = max(maxvallen, len(str(ssub_v)))
                sub_v = count_fn(sub)
                maxvallen = max(maxvallen, len(str(sub_v)))
                out.write(f"{sub_v:{maxvallen}d}\n")
            out.write(f'{"Total":{maxkeylen}.{maxkeylen}s} ')
            out.write(
                " ".join(
                    f"{tally:{maxvallen}d}" for k, tally in sorted(keytally.items())
                )
            )
            out.write(f" {sum(tally for k, tally in keytally.items()):{maxvallen}d}\n")
        else:
            raise ValueError("can only dump summary counts for 1 or 2-attribute pivots")

    def as_table(
            self,
            fn: Optional[Callable] = None,
            col: Optional[str] = None,
            col_label: Optional[str] = None
    ) -> Table:
        """Dump out the summary counts of this pivot table as a Table."""
        if col_label is None:
            col_label = col
        if fn is None:
            fn = len
            if col_label is None:
                col_label = "count"
        ret: Table = Table()
        col_label = cast(str, col_label)

        for attr in self._pivot_attrs:
            ret.create_index(attr)
        if len(self._pivot_attrs) == 1:
            for sub in self.subtables:
                subattr, subval = sub._attr_path[-1]
                attrdict = {subattr: subval}
                if col is None or fn is len:
                    attrdict[col_label] = fn(sub)
                else:
                    attrdict[col_label] = fn([getattr(s, col, None) for s in sub])
                ret.insert(default_row_class(**attrdict))
        elif len(self._pivot_attrs) == 2:
            for sub in self.subtables:
                for ssub in sub.subtables:
                    attrdict = dict(ssub._attr_path)
                    if col is None or fn is len:
                        attrdict[col_label] = fn(ssub)
                    else:
                        attrdict[col_label] = fn([getattr(s, col, None) for s in ssub])
                    ret.insert(default_row_class(**attrdict))
        elif len(self._pivot_attrs) == 3:
            for sub in self.subtables:
                for ssub in sub.subtables:
                    for sssub in ssub.subtables:
                        attrdict = dict(sssub._attr_path)
                        if col is None or fn is len:
                            attrdict[col_label] = fn(sssub)
                        else:
                            attrdict[col_label] = fn(
                                [getattr(s, col, None) for s in sssub]
                            )
                        ret.insert(default_row_class(**attrdict))
        else:
            raise ValueError("can only dump summary counts for 1 or 2-attribute pivots")
        return ret

    summary_counts = as_table

    def summarize(
            self,
            count_fn: Callable[[Iterable], int] = len,
            col_label: Optional[str] = None
    ) -> _PivotTableSummary:
        if col_label is None:
            if len(self._pivot_attrs) == 1:
                col_label = self._pivot_attrs[0]
            else:
                col_label = "value"
        return _PivotTableSummary(self, self._pivot_attrs, count_fn, col_label)


class _PivotTableSummary:
    def __init__(
            self,
            pivot_table: _PivotTable,
            pivot_attrs: list[str],
            count_fn: Callable[[Iterable], int] = len,
            col_label: Optional[str] = None
    ):
        self._pt = pivot_table
        self._pivot_attrs = pivot_attrs
        self._fn = count_fn
        self._label = col_label

    def as_html(self, **kwargs):
        formats = kwargs.get("formats", {})
        if len(self._pivot_attrs) == 1:
            col = self._pivot_attrs[0]
            col_label = self._label
            data: Table = Table().insert_many(
                default_row_class(**{col: k, col_label: self._fn(sub)})
                for k, sub in self._pt.items()
            )
            return data.as_html((col, col_label), formats=formats)

        elif len(self._pivot_attrs) == 2:
            keytally = dict.fromkeys(self._pt.subtables[0].keys(), 0)
            hdgs = [self._pivot_attrs[0]] + sorted(keytally) + ["Total"]

            def row_to_tr(r):
                ret_tr = ["<tr>"]
                for v, hdg in zip(r, hdgs):
                    v_format = formats.get(hdg, formats.get(type(v), "{}"))
                    v_align = "right" if isinstance(v, right_justify_types) else "left"
                    str_v = (
                        v_format.format(v) if isinstance(v_format, str) else v_format(v)
                    )
                    ret_tr.append(
                        f'<td><div align="{v_align}">{str_v}</div></td>'
                    )
                ret_tr.append("</tr>\n")
                return "".join(ret_tr)

            ret = ""
            ret += "<table>\n"
            ret += "<thead>\n"
            keytally = dict.fromkeys(self._pt.subtables[0].keys(), 0)
            hdgs = sorted(keytally)
            ret += (
                "<tr>"
                + "".join(f'<th><div align="center">{h}</div></th>' for h in hdgs)
                + "</tr>\n"
            )
            ret += "</thead>\n<tbody>\n"

            for k, sub in self._pt.items():
                row = [k]
                ssub_v_accum = 0
                for kk, ssub in sub.items():
                    ssub_v = self._fn(ssub)
                    row.append(ssub_v)
                    keytally[kk] += ssub_v
                    ssub_v_accum += ssub_v
                sub_v = ssub_v_accum  # count_fn(sub)
                row.append(sub_v)
                ret += row_to_tr(row)
            row: list[str | int] = ["Total"]
            row.extend(v for k, v in sorted(keytally.items()))
            row.append(sum(keytally.values()))
            ret += row_to_tr(row)

            ret += "</tbody>\n</table>\n"
            return ret

        else:  # if len(self._pivot_attrs) >= 3:
            raise Exception(
                "no HTML output format for 3-attribute pivot tables at this time"
            )


class _JoinTerm:
    """
    Temporary object created while composing a join across tables using
    L{Table.join_on} and '+' addition. JoinTerm's are usually created by
    calling join_on on a Table object, as in::

        customers.join_on("id") + orders.join_on("custid")

    This join expression would set up the join relationship
    equivalent to::

        customers.join(orders, id="custid")

    If tables are being joined on attributes that have the same name in
    both tables, then a join expression could be created by adding a
    JoinTerm of one table directly to the other table::

        customers.join_on("custid") + orders

    Once the join expression is composed, the actual join is performed
    using function call notation::

        customerorders = customers.join_on("custid") + orders
        for custord in customerorders():
            print custord

    When calling the join expression, you can optionally specify a
    list of attributes as defined in L{Table.join}.
    """

    def __init__(self, source_table: Table, join_field: str, join_type=None):
        self.source_table = source_table
        self.join_field = join_field
        self.join_to = None
        self.join_type = join_type

    def __add__(self, other: Union[Table, _JoinTerm]) -> _JoinTerm:
        if isinstance(other, Table):
            other = other.join_on(self.join_field)
        if isinstance(other, _JoinTerm):
            if self.join_to is None:
                if other.join_to is None:
                    self.join_to = other
                else:
                    self.join_to = other()

                if self.join_type is None:
                    self.join_type = other.join_type
                if other.join_type is None:
                    other.join_type = self.join_type

                return self
            else:
                if other.join_to is None:
                    return self() + other
                else:
                    return self() + other()
        raise ValueError(
            f"cannot add object of type {type(other).__name__!r} to JoinTerm"
        )

    def __radd__(self, other: Table) -> _JoinTerm:
        if isinstance(other, Table):
            return other.join_on(self.join_field) + self
        raise ValueError(
            f"cannot add object of type {type(other).__name__!r} to JoinTerm"
        )

    def __call__(self, attrs: Optional[Iterable[str]] = None) -> Table:
        if self.join_to:
            other = self.join_to
            if isinstance(other, Table):
                other = other.join_on(self.join_field)
            ret = self.source_table.join(
                other.source_table, attrs, **{self.join_field: other.join_field}
            )
            return ret
        else:
            return self.source_table.clone()

    def join_on(self, col: str) -> _JoinTerm:
        return self().join_on(col)


# Mixin classes for structure types that don't implement __eq__ sufficiently
try:
    from traits import HasTraits
except ImportError:
    try:
        from traitlets import HasTraits
    except ImportError:
        HasTraits = None

if HasTraits is not None:
    class HasTraitsMixin(HasTraits):
        def __eq__(self, other):
            return (isinstance(other, type(self)) and
                    all(getattr(self, attr) == getattr(other, attr) for attr in self.trait_names()))


if __name__ == "__main__":

    json_dumps = partial(json.dumps, indent=2)

    rawdata = textwrap.dedent(
        """\
        city,state,frequency,band,stn
        SANTA BARBARA,CA,1290,AM,KZSB
        CAVE CREEK,AZ,1100,AM,KFNX
        HONOLULU,HI,1270,AM,KNDI
        YOUNGSTOWN,OH,1390,AM,WNIO
        FREDERIKSTED,VI,1620,AM,WDHP
        GALVESTON,TX,1540,AM,KGBC
        SEASIDE,OR,840,AM,KSWB
        LILLINGTON,NC,1370,AM,WLLN
        POULTNEY,VT,1340,AM,WVNR
        GILLETTE,WY,1270,AM,KIML
        LITHIA SPRINGS,GA,890,AM,WJTP
        DICKINSON,ND,1230,AM,KDIX
        ABERDEEN,SD,1560,AM,KKAA
        ABERDEEN,SD,930,AM,KSDN
        ABERDEEN,WA,1320,AM,KXRO
        ABERDEEN,MS,1240,AM,WWZQ
        ABERDEEN,SD,1420,AM,KGIM
        ABERDEEN,WA,1450,AM,KBKW
        ABERDEEN,WA,1490,AM,KWOK
        """
    )

    # load miniDB
    stations = Table().csv_import(rawdata, transforms={"frequency": int})

    # create unique index by station call letters
    stations.create_index("stn", unique=True)

    # perform some queries and deletes
    queries: list[dict] = [
        {"city": "ABERDEEN"},
        {"city": "ABERDEEN", "stn": "KGIM"},
        {"state": "WA"},
        {"frequency": Table.gt(1000)},
    ]
    for query_args in queries:
        print(f"Query radio stations: {query_args}")
        result = stations.where(**query_args)
        print(len(result))
        for rec in result:
            print(rec)
        print()

    print(stations.delete(city="ABERDEEN"))
    print(list(stations.where()))
    print()

    amfm: Table = Table()
    amfm.create_index("stn", unique=True)
    amfm.insert({'stn': "KFNX", 'band': "AM"})
    amfm.insert({'stn': "KPHX", 'band': "FM"})
    amfm.insert({'stn': "KPHA", 'band': "FM"})
    amfm.insert({'stn': "KDFW", 'band': "FM"})
    print(amfm.by.stn["KFNX"])
    print(amfm.by.stn["KFNX"].band)

    try:
        amfm.insert({'stn': "KPHA", 'band': "AM"})
    except KeyError:
        print("duplicate key not allowed")

    print()
    for rec in (stations.join_on("stn") + amfm.join_on("stn"))(
        ["stn", "city", (amfm, "band", "AMFM"), (stations, "state", "st")]
    ).orderby("AMFM"):
        print(repr(rec))

    print()
    for rec in (stations.join_on("stn") + amfm.join_on("stn"))(
        ["stn", "city", (amfm, "band"), (stations, "state", "st")]
    ):
        print(json_dumps(vars(rec)))

    print()
    for rec in (stations.join_on("stn") + amfm.join_on("stn"))():
        print(json_dumps(vars(rec)))

    print()
    stations.create_index("state")
    for az_stn in stations.by.state["AZ"]:
        print(az_stn)

    print()
    pivot = stations.pivot("state")
    pivot.dump_counts()

    print()
    stations.create_index("band")
    pivot = (stations.join_on("stn") + amfm)().pivot("state band")
    pivot.dump_counts()

    print()
    for rec in amfm:
        print(rec)
    print()

    print(list(amfm.all.stn))
    print(list(amfm.all.band))
    print(list(amfm.unique("band").all.band))
    print(list(amfm.all.band.unique))
    print()

    del amfm[0:-1:2]
    for stn in amfm:
        print(stn)

    print()
    print(amfm.pop(-1))
    print(len(amfm))
    print(amfm.by.stn["KPHX"])
    try:
        print(amfm.by.stn["KPHY"])
    except KeyError:
        print("no station 'KPHY' in table")

    print(list(stations.all.stn))

    # do some simple stats with common ML data set
    url = "https://raw.githubusercontent.com/jbrownlee/Datasets/master/iris.csv"
    names = ["sepal-length", "sepal-width", "petal-length", "petal-width", "class"]
    iris_transforms = dict.fromkeys(
        ["petal-length", "petal-width", "sepal-length", "sepal-width"], float
    )
    iris_table = Table("iris").csv_import(
        url, fieldnames=names, transforms=iris_transforms
    )

    print(iris_table.info())
    for rec in iris_table[:5]:
        print(rec)

    stats = iris_table.stats(
        ["petal-length", "petal-width", "sepal-length", "sepal-width"]
    )
    for rec in stats:
        print(rec)
